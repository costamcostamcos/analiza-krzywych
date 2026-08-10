# =====================================================================
# INTERAKTYWNY ANALIZATOR KRZYWYCH AI PRO — WERSJA 5
# =====================================================================
# NOWE W WERSJI 5:
#     [v5.0] Zarządzanie pamięcią RAM (odpowiedź na throttling Streamlit):
#     wszystkie 5 buforów @st.cache_data dostało limity max_entries oraz ttl,
#     więc cache nie rośnie już w nieskończoność przy zmianach parametrów
#     (to była główna przyczyna narastania RAM aż do rebootów). Limity:
#     silnik klastrowania 60 wpisów/1h (wołany wielokrotnie przez ranking,
#     ale wynik mały), metryki K 10/1h, Google Sheets 3/30 min, LOO 3/30 min,
#     ranking 3/30 min (ciężkie, duże wyniki — trzymamy tylko kilka ostatnich).
#     Dodano przycisk „Wyczyść pamięć podręczną” w panelu bocznym
#     (st.cache_data.clear() + gc.collect()).
#     Eksporty (CSV/Excel) mają teraz opisowe nazwy plików zawierające: nazwę
#     pliku źródłowego, metodę, obróbkę wstępną i liczbę grup K (a przy
#     klasyfikacji nowych widm także metodę przypisania i nazwę pliku nowych
#     widm). Ułatwia to rozróżnienie wielu eksportów o podobnej treści.
#
# NOWE W WERSJI 4:
#  D. Klasyfikacja regułowa widm EPR wg drzew decyzyjnych z pracy
#     Marciniak et al. (2025), Front. Public Health 13:1659601 (Fig. 1A/1B).
#     Deterministyczne przypisanie widma do jednego z 5 typów line-shape
#     (I–V, z podziałem IVA/IVB) na podstawie znaku sygnału w wybranych
#     wartościach g oraz obecności lokalnych ekstremów. W pełni
#     interpretowalna (zwraca ścieżkę decyzyjną), bez uczenia. Działa na
#     tej samej ujednoliconej osi X co reszta aplikacji. Nowa sekcja
#     rozwijana z tabelą typów, podglądem widm z punktami decyzyjnymi,
#     eksportem CSV oraz — gdy dostępny Ground Truth — macierzą zgodności
#     etykiet regułowych z podziałem eksperckim.
#     [v4.1] Wygładzanie do wyboru: średnia ruchoma LUB Savitzky-Golay
#     (zachowuje wysokość pików — bliżej praktyki EPR). Detekcja ekstremów
#     do wyboru: prominencja amplitudowa LUB zmiana znaku pochodnej. Próg
#     prominencji jest w bezwzględnych jednostkach sygnału (a.u.); obok
#     wyświetlana jest szacowana σ szumu linii bazowej jako punkt odniesienia
#     (sugerowany próg ≈ 3σ) i zarazem miara jakości widma.
#     [v4.2] Ranking skuteczności ocenia teraz STABILNOŚĆ: każda metoda jest
#     uruchamiana na N ziarnach losowych, a wyniki podawane jako średnia ± σ.
#     Metody deterministyczne (Ward, korelacyjna, PCA+Ward) mają σ≈0. Wykres
#     porównawczy ma słupki błędów, tabela — kolumny σ i „Rozrzut σ", a
#     podsumowanie wskazuje osobno metodę najskuteczniejszą i najstabilniejszą.
#     Cel: odróżnić realną przewagę metody od szczęśliwego trafienia ziarna
#     na małym zbiorze.
#     [v4.3] Odporne wczytywanie plików: błędy formuł Excela (#REF!, #DIV/0!,
#     #VALUE! itp.) są traktowane jak puste komórki (kolumna częściowo zepsuta
#     jest pomijana, nie wywala całości). Gdy po oczyszczeniu nie ma danych
#     liczbowych, użytkownik dostaje czytelny komunikat po polsku zamiast
#     kryptycznego IndexError (nowy wyjątek BladDanychWejsciowych).
#     [v4.4] Naprawiono segmentation fault (crash OOM) przy rankingu w trybie
#     „Pełny". Ranking wielo-ziarnowy dobiera teraz liczbę ziaren per metoda:
#     deterministyczne liczą się RAZ (σ=0), ciężkie (UMAP/Spectral/SOM/Konsensus)
#     mają limit 3 ziaren, lekkie losowe pełne N. Pamięć zwalniana (gc) po każdej
#     metodzie. Efekt: przy 17 metodach × N=5 liczba uruchomień silnika spadła
#     z 85 do 43 (ciężkie z 40 do 18) — bez utraty wartości oceny stabilności.
#     [v4.5] Oś g-factor rysowana MALEJĄCO w prawo na wszystkich wykresach widm
#     (konwencja EPR — pole magnetyczne rośnie w prawo). Dotyczy wykresów w
#     aplikacji, obrazków PNG oraz natywnych wykresów w eksporcie do Excela.
#     To zmiana wyłącznie wizualna — dane i obliczenia (klasyfikacja regułowa,
#     klastrowanie, interpolacja) nadal działają na rosnącej siatce g. Wykresy
#     bez osi g (sugestia K, porównanie metod) pozostają bez zmian.
#     [v4.6] Porządki: przestarzały parametr use_container_width=True zastąpiony
#     przez width='stretch' we wszystkich 26 miejscach (wykresy, tabele, przyciski)
#     — usuwa ostrzeżenia deprecation z logów i zabezpiecza przed przyszłym
#     usunięciem parametru przez Streamlit. Zamianę przetestowano na wersji 1.59.1.
#     Usunięto też nieużywany tooltip sidebara (wstrzyknięcie JS przez
#     components.html) — nie działał, generował ostrzeżenie deprecation i sięgał
#     do window.parent.document (kruche). Stylizacja przycisku sidebara (pasek
#     „Grupy Wzorcowe") została — to czysty CSS, bez JS. Import components usunięty.
#
# NOWE W WERSJI 3:
#  A. Ujednolicenie osi X: widma wzorcowe i nowe przechodzą przez TĘ SAMĄ
#     funkcję (sortowanie rosnące + interpolacja na wspólną siatkę). Dzięki
#     temu wgranie tego samego pliku jako „nowe” daje self-match = 0, a k-NN
#     wiernie odtwarza podział ekspercki. Dodano listę „uciekinierów”
#     (widm, których przypisanie różni się od etykiety eksperckiej).
#  B. Nazwy klastrów literami grup eksperckich (a → „Klaster A”) przez
#     OPTYMALNE przypisanie (algorytm węgierski) + procent czystości.
#     Klastry nadmiarowe: „Klaster N (mieszany)”. Bez Ground Truth pozostają
#     nazwy numeryczne.
#  C. Eksport wykresów do Excela: dane źródłowe w postaci kolumn X + serie
#     Y (arkusze „Dane — krzywe”, „Dane — profile”, „Dane — wykres”) do
#     samodzielnej edycji (kolory, typ linii, dodawanie/usuwanie serii),
#     obok podglądu PNG 300 DPI. Dodatkowo natywny, edytowalny wykres
#     Excela odtwarzający styl aplikacji (profile wzorcowe — linie ciągłe,
#     nowe widma — linie przerywane w kolorze przypisanej kategorii).
#     Klasyfikacja nowych widm ma teraz eksport do Excela (tabela przypisań
#     + dane wykresu nakładkowego + natywny wykres + podgląd).
#
# CHANGELOG względem wersji pierwotnej:
#
# [BŁĘDY]
#  1. Naprawiono martwy przycisk CSV w sidebarze (klucz "df_eksport_klastry"
#     nie był nigdy zapisywany; teraz przycisk dodawany jest po obliczeniach).
#  2. "Spectral + GMM" faktycznie używa embeddingu spektralnego przed GMM.
#  3. "Spectral + Hierarchiczna" używa embeddingu spektralnego zamiast
#     wierszy macierzy afiniczności.
#  4. SiecSOM to teraz prawdziwa mapa Kohonena: gaussowska funkcja
#     sąsiedztwa + malejący promień sigma (wcześniej: tylko BMU = online
#     k-means bez topologii).
#  5. Time Warping przepisany na spójne, monotoniczne odkształcenie osi
#     czasu (wcześniej mapowanie węzłów było niespójne).
#  6. Ranking przy "Augmentacji sygnału" porównuje z ekspertem tylko
#     oryginalne krzywe (wcześniej cichy wyjątek usuwał metody z rankingu).
#  7. Cache silnika klastrowania hashuje też surowe dane (usunięty
#     podkreślnik przy df_sygnaly_raw) — koniec ryzyka stęchłego cache'a.
#
# [METODOLOGIA]
#  8. ARI/NMI liczone TYLKO gdy istnieje realny Ground Truth (arkusz w
#     pliku). Domyślny sztywny podział y1–y43 stosowany wyłącznie, gdy
#     nazwy krzywych do niego pasują — z wyraźnym ostrzeżeniem w UI.
#     Bez GT metryki zewnętrzne są ukrywane zamiast liczone od fikcji.
#
# [WYDAJNOŚĆ]
#  9. Leave-One-Out uruchamiany dopiero po kliknięciu przycisku i cache'owany
#     (wcześniej ~43 pełne klasteryzacje przy KAŻDYM rerunie).
# 10. Klastrowanie konsensusowe zwektoryzowane: O(4·N²) w NumPy zamiast
#     potrójnej pętli w Pythonie.
# 11. Standaryzacja krzywych liczona raz i współdzielona (sugestia K,
#     MSE, LOO). Google Sheets pobierany jednym requestem (sheet_name=None).
#
# [PORZĄDKI]
# 12. Stałe konfiguracyjne zebrane w sekcji KONFIGURACJA.
# 13. Usunięte duplikaty (PLOTLY_KOLORY x2, importy metryk x2, martwy
#     import PyTorch), zmienna `górna` -> `gorna`.
# 14. Walidacja rozmiaru danych (zakres K ograniczony liczbą krzywych,
#     ochrona silhouette/elbow przy małych zbiorach).
# 15. Podgląd wczytanej tabeli, żeby heurystyka nagłówka była widoczna.
# 16. Tryb debug: pełny traceback zamiast jednolinijkowego błędu.
#
# [WERSJA 2.1]
# 17. Metody czysto hierarchiczne (dendrogramowe, np. Ward) pokazują teraz
#     także Wykres 2 (profile modelowe), szczegółowy skład klastrów oraz
#     eksport CSV/Excel — wcześniej sekcje te były dla nich ukryte, mimo że
#     przypisania z fcluster() były liczone (służyły do ARI/NMI).
#
# [WERSJA 2.2]
# 18. Nowa sekcja "Klasyfikacja nowych widm": drugi uploader przyjmuje
#     skoroszyt Excel z nowymi widmami (ten sam układ: X + kolumny widm),
#     interpoluje je na siatkę referencyjną, standaryzuje TYM SAMYM
#     skalerem co dane wzorcowe i przypisuje do kategorii eksperckich
#     (k-NN ważone odległością lub najbliższy centroid), z pewnością
#     przypisania, wykresem nakładkowym i eksportem CSV. Bez Ground Truth
#     klasyfikacja odbywa się do klastrów z bieżącej analizy.
#
# [WERSJA 2.3]
# 19. NAPRAWIONO płaskie linie nowych widm: np.interp wymaga rosnącej osi X,
#     a widma EPR mają ją często malejącą (od wysokiego pola do niskiego) lub
#     nieposortowaną — bez sortowania interpolacja zwracała wartości stałe
#     (płaskie kreski) zamiast kształtu widma. Teraz oś X nowych i
#     referencyjnych widm jest sortowana rosnąco przed interpolacją, z
#     ostrzeżeniem, gdy zakresy X się nie pokrywają.
#
# [WERSJA 2.4]
# 20. Trzecia metoda oceny przypisania: "Skalibrowane prawdopodobieństwo"
#     (SVM RBF + kalibracja Platta z walidacją krzyżową, CalibratedClassifierCV).
#     W odróżnieniu od k-NN (zgodność głosów) i centroidu (margines separacji),
#     zwraca przybliżenie rzeczywistego P(kategoria|widmo). Liczba foldów CV
#     dobierana automatycznie do najmniejszej kategorii; ostrzeżenie przy
#     małych zbiorach. Kolumna pewności i jej opis są teraz zależne od metody
#     (uczciwe nazwy: "Prawdopodobieństwo (%)", "Zgodność głosów (%)",
#     "Margines separacji") zamiast mylącego wspólnego "Pewność (%)".
# =====================================================================

import io
import gc

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import plotly.graph_objects as go
from plotly.subplots import make_subplots

import streamlit as st

from sklearn.preprocessing import StandardScaler, MinMaxScaler
from sklearn.cluster import KMeans, HDBSCAN, SpectralClustering
from sklearn.mixture import GaussianMixture, BayesianGaussianMixture
from sklearn.decomposition import NMF, PCA
from sklearn.manifold import SpectralEmbedding
from sklearn.neighbors import KNeighborsClassifier, NearestCentroid
from sklearn.calibration import CalibratedClassifierCV
from sklearn.svm import SVC
from collections import Counter
from sklearn.metrics import (
    silhouette_score,
    adjusted_rand_score,
    normalized_mutual_info_score,
    davies_bouldin_score,
    calinski_harabasz_score,
)
from scipy.cluster.hierarchy import dendrogram, linkage, fcluster
from scipy.spatial.distance import squareform
from scipy.optimize import linear_sum_assignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties

# --- Bezpieczne importy opcjonalnych zależności ---
try:
    import umap
    umap_dostepne = True
except ImportError:
    umap_dostepne = False

try:
    from tslearn.clustering import KShape
    from tslearn.utils import to_time_series_dataset
    tslearn_dostepne = True
except ImportError:
    tslearn_dostepne = False


# =====================================================================
# KONFIGURACJA — wszystkie "magiczne liczby" w jednym miejscu
# =====================================================================

KONFIG = {
    "ROLLING_WINDOW": 5,           # okno filtrowania szumów (rolling mean)
    "SOM_X": 5,                    # wymiary siatki Kohonena
    "SOM_Y": 5,
    "SOM_EPOKI": 50,
    "SOM_LR": 0.5,
    "UMAP_N_NEIGHBORS": 15,
    "UMAP_MIN_DIST": 0.05,
    "UMAP_MIN_DIST_PREPROC": 0.1,
    "K_MAX_SUGESTIA": 10,          # górna granica przeszukiwania K
    "PCA_KOMPONENTY": 3,
    "RANDOM_STATE": 42,
    "DEBUG": False,                # True = pełny traceback przy błędzie
}

PLOTLY_KOLORY = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
]

NAZWY_KOLOROW = [
    "Niebieski", "Pomarańczowy", "Zielony", "Czerwony", "Fioletowy",
    "Brązowy", "Różowy", "Szary", "Oliwkowy", "Jasnoniebieski",
]


# =====================================================================
# NAZWY PLIKÓW EKSPORTU — opisowe, żeby po pobraniu kilku plików było
# jasne, jakich ustawień i danych dotyczą. Fragmenty: nazwa pliku
# źródłowego, metoda, obróbka wstępna, liczba grup K.
# =====================================================================
def _oczysc_do_nazwy(tekst, maxlen=40):
    """Zamienia dowolny tekst na bezpieczny fragment nazwy pliku:
    bez rozszerzenia, bez spacji/znaków specjalnych, skrócony."""
    if not tekst:
        return ""
    tekst = str(tekst)
    # odetnij rozszerzenie pliku (.xlsx, .csv) jeśli jest
    for ext in (".xlsx", ".xls", ".csv", ".XLSX", ".XLS", ".CSV"):
        if tekst.endswith(ext):
            tekst = tekst[: -len(ext)]
            break
    # spolszczone znaki -> ascii-podobne, reszta -> _
    zamiany = {
        "ą": "a", "ć": "c", "ę": "e", "ł": "l", "ń": "n", "ó": "o",
        "ś": "s", "ź": "z", "ż": "z", "Ą": "A", "Ć": "C", "Ę": "E",
        "Ł": "L", "Ń": "N", "Ó": "O", "Ś": "S", "Ź": "Z", "Ż": "Z",
    }
    for zle, dobre in zamiany.items():
        tekst = tekst.replace(zle, dobre)
    dozwolone = []
    for znak in tekst:
        if znak.isalnum() or znak in ("-", "_"):
            dozwolone.append(znak)
        elif znak in (" ", ".", ",", "/", "\\", "+", "(", ")"):
            dozwolone.append("_")
    wynik = "".join(dozwolone).strip("_")
    # zredukuj wielokrotne podkreślenia
    while "__" in wynik:
        wynik = wynik.replace("__", "_")
    return wynik[:maxlen]


def zbuduj_nazwe_eksportu(prefiks, rozszerzenie, nazwa_pliku=None,
                          metoda=None, obrobka=None, k=None, dodatek=None):
    """Buduje opisową nazwę pliku eksportu.

    Przykład: 'klastry__widma_probki__K-means__Standardowa__K3.csv'
    Puste elementy są pomijane, więc funkcja działa też przy braku danych.
    """
    czesci = [prefiks]
    if nazwa_pliku:
        czesci.append(_oczysc_do_nazwy(nazwa_pliku))
    if metoda:
        czesci.append(_oczysc_do_nazwy(metoda, maxlen=25))
    if obrobka:
        czesci.append(_oczysc_do_nazwy(obrobka, maxlen=20))
    if k is not None:
        czesci.append(f"K{k}")
    if dodatek:
        czesci.append(_oczysc_do_nazwy(dodatek, maxlen=25))
    czesci = [c for c in czesci if c]
    return "__".join(czesci) + rozszerzenie


# =====================================================================
# UJEDNOLICENIE OSI X (wspólna siatka rosnąca)
# Ta sama transformacja stosowana do widm wzorcowych ORAZ nowych,
# aby identyczne wejście dawało identyczną reprezentację liczbową.
# Zwraca: (x_ref_rosnace, funkcja_przygotuj(macierz_widm, x_widm)).
# =====================================================================
def zbuduj_ujednolicacz_osi(x_ref_raw):
    """Buduje wspólny „ujednolicacz” osi X w oparciu o oś referencyjną.

    Oś referencyjna jest sortowana rosnąco (widma EPR często mają oś
    malejącą). Zwrócona funkcja przygotowuje DOWOLNĄ macierz widm
    (wiersze = punkty osi X, kolumny = widma) na tę samą rosnącą siatkę:
    sortuje wejściową oś, interpoluje liniowo na siatkę referencyjną i —
    jeśli osie są już identyczne — zwraca dane bez zmian (bez interpolacji).
    """
    x_ref_raw = np.asarray(x_ref_raw, dtype=float)
    if x_ref_raw[0] > x_ref_raw[-1]:
        kol_ref = np.argsort(x_ref_raw)
        x_ref = x_ref_raw[kol_ref]
    else:
        kol_ref = np.arange(len(x_ref_raw))
        x_ref = x_ref_raw

    def przygotuj(macierz_widm, x_widm):
        """macierz_widm: (n_punktow, n_widm); x_widm: (n_punktow,).
        Zwraca (n_widm, n_punktow) na rosnącej siatce referencyjnej."""
        macierz_widm = np.asarray(macierz_widm, dtype=float)
        x_widm = np.asarray(x_widm, dtype=float)

        # Ścieżka „bez interpolacji”: identyczna oś (po ewentualnym
        # posortowaniu referencyjnej) → zero perturbacji numerycznych.
        if len(x_widm) == len(x_ref) and np.allclose(x_widm, x_ref):
            return macierz_widm.T.copy()

        # Jeśli to dokładnie oś referencyjna w oryginalnej (malejącej)
        # kolejności — wystarczy przełożyć wiersze wg kol_ref, bez interpolacji.
        if len(x_widm) == len(x_ref_raw) and np.allclose(x_widm, x_ref_raw):
            return macierz_widm[kol_ref, :].T.copy()

        # W pozostałych przypadkach: sortujemy oś wejściową i interpolujemy.
        kolejnosc = np.argsort(x_widm)
        x_sort = x_widm[kolejnosc]
        wynik = np.column_stack([
            np.interp(x_ref, x_sort, macierz_widm[:, j][kolejnosc])
            for j in range(macierz_widm.shape[1])
        ])
        return wynik.T.copy()

    return x_ref, przygotuj


# =====================================================================
# KLASYFIKACJA REGUŁOWA WIDM EPR (Marciniak et al. 2025, Fig. 1A/1B)
# =====================================================================
# Deterministyczne drzewa decyzyjne z pracy „Categorization of screen
# glasses of mobile devices...”. Klasyfikacja opiera się na znaku sygnału
# w wybranych wartościach g oraz obecności lokalnych ekstremów w zadanych
# zakresach g. Nie wymaga uczenia — jest w pełni interpretowalna i zwraca
# ścieżkę decyzyjną. Działa na widmach po ujednoliceniu osi X (ta sama
# siatka co reszta aplikacji), więc znaki f(g) są liczone spójnie.
#
# UWAGA metodologiczna: progi znakowe zakładają preprocessing zbliżony do
# pracy (odjęcie tła rurki, liniowa korekcja bazy, normalizacja). Dla
# innych spektrometrów/parametrów akwizycji może być konieczne dostrojenie
# okna wygładzania i prominencji ekstremów.

# Punkty i zakresy g używane przez drzewa decyzyjne (z Fig. 1)
_REG_PUNKTY_G = [2.0000, 2.0043, 2.0171]
_REG_ZAKRESY_G = {
    "min_III_V": (2.0001, 2.0040),   # lokalne minimum -> V, brak -> III
    "max_II": (2.0200, 2.0250),      # lokalne maksimum -> II
    "min_I_IVA": (1.9930, 1.9990),   # lokalne minimum -> I, brak -> IVA
}


def _reg_sort(g, y):
    g = np.asarray(g, dtype=float)
    y = np.asarray(y, dtype=float)
    idx = np.argsort(g)
    return g[idx], y[idx]


def _reg_smooth(y, window=7, metoda="Średnia ruchoma"):
    """Wygładza sygnał wybraną metodą.

    - „Średnia ruchoma": prosty filtr pudełkowy (splot). Szybki, ale
      spłaszcza i przesuwa piki — może zaniżać amplitudę ekstremów.
    - „Savitzky-Golay": dopasowuje lokalnie wielomian (rząd 3). Zachowuje
      wysokość i położenie pików znacznie lepiej — bliżej praktyki EPR.
    """
    if window is None or window < 3:
        return np.asarray(y, dtype=float)
    window = window + 1 if window % 2 == 0 else window
    y = np.asarray(y, dtype=float)
    if window > len(y):
        window = len(y) - 1 if len(y) % 2 == 0 else len(y)
        if window < 3:
            return y
    if metoda == "Savitzky-Golay":
        try:
            from scipy.signal import savgol_filter
            polyorder = min(3, window - 1)
            return savgol_filter(y, window_length=window, polyorder=polyorder)
        except Exception:
            pass  # awaryjnie: średnia ruchoma
    kernel = np.ones(window) / window
    return np.convolve(y, kernel, mode="same")


def reg_szum_sigma(g, y, g_lo=2.024, g_hi=2.030, smooth_window=7,
                   metoda_smooth="Średnia ruchoma"):
    """Szacuje odchylenie standardowe szumu linii bazowej.

    Bierze fragment widma z dala od głównych sygnałów (domyślnie skrzydło
    g ∈ [2.024; 2.030]), odejmuje jego wygładzoną wersję (usuwa ewentualny
    trend/dryf bazy) i liczy σ rezyduów. To orientacyjna miara szumu — do
    doboru progu prominencji (np. 3σ). Zwraca float albo None, gdy zakres
    ma za mało punktów.
    """
    gs, ys = _reg_sort(g, y)
    lo, hi = min(g_lo, g_hi), max(g_lo, g_hi)
    maska = (gs >= lo) & (gs <= hi)
    if maska.sum() < 5:
        return None
    seg = ys[maska]
    baza = _reg_smooth(seg, min(smooth_window, len(seg)), metoda_smooth)
    return float(np.std(seg - baza))


def reg_value_at_g(g, y, g0):
    """Interpolowana wartość sygnału f(g0)."""
    gs, ys = _reg_sort(g, y)
    return float(np.interp(g0, gs, ys))


def _reg_extremum(g, y, g_lo, g_hi, kind="min", smooth_window=7,
                  prominence=0.0, metoda_smooth="Średnia ruchoma",
                  metoda_ekstr="Prominencja amplitudowa"):
    """Czy w zakresie g ∈ (g_lo; g_hi) istnieje lokalne ekstremum danego typu?

    Dwie metody wykrywania:
    - „Prominencja amplitudowa": ekstremum leży wewnątrz okna (nie na brzegu)
      i wystaje ponad wartości brzegowe o co najmniej `prominence`. Prosta,
      progowana amplitudowo.
    - „Zmiana znaku pochodnej": ekstremum = punkt, w którym pierwsza pochodna
      wygładzonego sygnału zmienia znak (klasyczny warunek min/max). `prominence`
      działa tu jako dodatkowy filtr amplitudy odsiewający ekstrema szumowe.
    """
    gs, ys = _reg_sort(g, y)
    lo, hi = min(g_lo, g_hi), max(g_lo, g_hi)
    maska = (gs >= lo) & (gs <= hi)
    if maska.sum() < 5:
        raise ValueError(f"Za mało punktów w zakresie g=({lo}; {hi}).")
    seg = _reg_smooth(ys, smooth_window, metoda_smooth)[maska]
    sygn = -seg if kind == "max" else seg  # zawsze szukamy MINIMUM w `sygn`

    if metoda_ekstr == "Zmiana znaku pochodnej":
        # Ekstremum tam, gdzie pochodna przechodzi z (-) na (+): dolina.
        d = np.diff(sygn)
        kandydaci = [k for k in range(1, len(sygn) - 1)
                     if d[k - 1] < 0 <= d[k]]
        for k in kandydaci:
            # Filtr amplitudy: dolina musi wystawać ponad brzegi okna
            if (min(sygn[0], sygn[-1]) - sygn[k]) >= prominence:
                return True
        return False

    # Domyślnie: prominencja amplitudowa (jak dotychczas)
    i = int(np.argmin(sygn))
    wewnatrz = 0 < i < len(sygn) - 1
    wystajace = (min(sygn[0], sygn[-1]) - sygn[i]) >= prominence
    return bool(wewnatrz and wystajace)


def reg_klasyfikuj_nienapromienione(g, y, smooth_window=7, prominence=0.0,
                                    metoda_smooth="Średnia ruchoma",
                                    metoda_ekstr="Prominencja amplitudowa"):
    """Drzewo Fig. 1A (próbki 0 Gy). Zwraca (typ, lista_kroków)."""
    sciezka = []
    kw = dict(smooth_window=smooth_window, prominence=prominence,
              metoda_smooth=metoda_smooth, metoda_ekstr=metoda_ekstr)
    f2000 = reg_value_at_g(g, y, 2.0000)
    sciezka.append(f"f(2.0000) = {f2000:.3e}")
    if f2000 < 0:
        ma_min = _reg_extremum(g, y, 2.0040, 2.0001, "min", **kw)
        sciezka.append(f"lokalne min w (2.0040; 2.0001): {ma_min}")
        return ("V" if ma_min else "III"), sciezka
    f20171 = reg_value_at_g(g, y, 2.0171)
    sciezka.append(f"f(2.0171) = {f20171:.3e}")
    if f20171 < 0:
        return "IVB", sciezka
    ma_max = _reg_extremum(g, y, 2.0250, 2.0200, "max", **kw)
    sciezka.append(f"lokalne max w (2.0250; 2.0200): {ma_max}")
    if ma_max:
        return "II", sciezka
    ma_min = _reg_extremum(g, y, 1.9990, 1.9930, "min", **kw)
    sciezka.append(f"lokalne min w (1.9990; 1.9930): {ma_min}")
    return ("I" if ma_min else "IVA"), sciezka


def reg_klasyfikuj_napromienione(g, y, smooth_window=7, prominence=0.0,
                                 metoda_smooth="Średnia ruchoma",
                                 metoda_ekstr="Prominencja amplitudowa"):
    """Drzewo Fig. 1B (próbki 10 Gy). IVA i IVB nierozróżnialne. Zwraca (typ, kroki)."""
    sciezka = []
    kw = dict(smooth_window=smooth_window, prominence=prominence,
              metoda_smooth=metoda_smooth, metoda_ekstr=metoda_ekstr)
    f2000 = reg_value_at_g(g, y, 2.0000)
    sciezka.append(f"f(2.0000) = {f2000:.3e}")
    if f2000 < 0:
        ma_min = _reg_extremum(g, y, 2.0040, 2.0001, "min", **kw)
        sciezka.append(f"lokalne min w (2.0040; 2.0001): {ma_min}")
        return ("V" if ma_min else "III"), sciezka
    f20171 = reg_value_at_g(g, y, 2.0171)
    sciezka.append(f"f(2.0171) = {f20171:.3e}")
    if f20171 < 0:
        return "IVA&IVB", sciezka
    f20043 = reg_value_at_g(g, y, 2.0043)
    sciezka.append(f"f(2.0043) = {f20043:.3e}")
    return ("I" if f20043 > 0 else "II"), sciezka


# =====================================================================
# MAPOWANIE KLASTRÓW NA LITERY GRUP EKSPERCKICH (algorytm węgierski)
# Klaster, w którym dominują krzywe z eksperckiej grupy „a”, otrzymuje
# nazwę „Klaster A” itd. Przypisanie jest OPTYMALNE (maksymalizuje
# łączną liczbę trafnie nazwanych krzywych) i wzajemnie jednoznaczne.
# =====================================================================
def mapuj_klastry_na_litery(numery_grup, etykiety_eksperta):
    """Zwraca słownik: nr_klastra -> (etykieta_wyswietlana, czystosc_%).

    - Litery pochodzą z rzeczywistych etykiet eksperckich (a -> „A”).
    - Przypisanie klaster→grupa liczone algorytmem węgierskim na macierzy
      kontyngencji (ile krzywych z grupy g wpadło do klastra k).
    - Czystość = udział krzywych z przypisanej grupy w danym klastrze.
    - Klastry bez pary (więcej klastrów niż grup) oraz szum (0) dostają
      nazwę „Klaster N (mieszany)”.
    """
    numery_grup = np.asarray(numery_grup)
    etykiety_eksperta = np.asarray([str(e) for e in etykiety_eksperta])

    klastry = sorted(int(k) for k in set(numery_grup) if int(k) > 0)
    grupy = sorted(set(etykiety_eksperta))
    wynik = {}

    if not klastry or not grupy:
        for k in sorted(int(v) for v in set(numery_grup)):
            wynik[k] = (("Szum / Odrzuty" if k == 0 else f"Klaster {k}"), None)
        return wynik

    # Macierz kontyngencji: wiersze = klastry, kolumny = grupy eksperckie
    idx_klastra = {k: i for i, k in enumerate(klastry)}
    idx_grupy = {g: j for j, g in enumerate(grupy)}
    M = np.zeros((len(klastry), len(grupy)), dtype=float)
    for kl, gr in zip(numery_grup, etykiety_eksperta):
        kl = int(kl)
        if kl > 0:
            M[idx_klastra[kl], idx_grupy[gr]] += 1

    # Algorytm węgierski maksymalizuje liczbę trafień (minus koszt)
    wiersze, kolumny = linear_sum_assignment(-M)
    przypisana_grupa = {}
    for r, c in zip(wiersze, kolumny):
        przypisana_grupa[klastry[r]] = grupy[c]

    liczności_klastra = {k: int((numery_grup == k).sum()) for k in klastry}
    uzyte_litery = {}
    for k in klastry:
        if k in przypisana_grupa:
            g = przypisana_grupa[k]
            litera = str(g).upper()
            n_z_grupy = int(M[idx_klastra[k], idx_grupy[g]])
            czystosc = 100.0 * n_z_grupy / max(liczności_klastra[k], 1)
            # Zabezpieczenie przed zdublowaną literą (teoretycznie niemożliwe
            # przy 1:1, ale gdy dwie grupy mają identyczną literę po .upper()).
            if litera in uzyte_litery:
                litera = f"{litera}·{k}"
            uzyte_litery[litera] = k
            wynik[k] = (f"Klaster {litera}", round(czystosc, 1))
        else:
            # Klaster nadmiarowy — brak pary z grupą ekspercką
            wynik[k] = (f"Klaster {k} (mieszany)", None)

    if 0 in set(int(v) for v in numery_grup):
        wynik[0] = ("Szum / Odrzuty", None)
    return wynik


# =====================================================================
# RENDER WYKRESÓW PUBLIKACYJNYCH (matplotlib -> PNG do Excela)
# Kaleido (Plotly -> PNG) bywa niedostępne na Streamlit Cloud, dlatego
# wersje do publikacji rysujemy matplotlibem: pewne, wysokie DPI,
# kontrola czcionek i rozmiaru zgodnie z wymogami czasopism.
# =====================================================================
def _png_krzywe(x, krzywe_df, numery_grup, kolory_hex, dpi=300):
    fig, ax = plt.subplots(figsize=(7.0, 4.2))
    dodane = set()
    for i, col in enumerate(krzywe_df.columns):
        kid = int(numery_grup[i])
        kolor = "#aaaaaa" if kid <= 0 else kolory_hex[(kid - 1) % 10]
        etykieta = "Szum" if kid <= 0 else f"Klaster {kid}"
        ax.plot(x, krzywe_df[col], color=kolor, linewidth=0.9, alpha=0.7,
                label=etykieta if kid not in dodane else None)
        dodane.add(kid)
    ax.set_xlabel("Oś X"); ax.set_ylabel("Sygnał")
    ax.legend(fontsize=7, framealpha=0.85)
    ax.grid(True, color="#e0e0e0", linewidth=0.5)
    ax.invert_xaxis()  # oś g-factor malejąca w prawo — konwencja EPR
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=dpi)
    plt.close(fig); buf.seek(0)
    return buf


def _png_profile(x, krzywe_df, numery_grup, kolory_hex, dpi=300):
    fig, ax = plt.subplots(figsize=(7.0, 4.2))
    for kid in sorted(set(int(v) for v in numery_grup)):
        maska = [int(numery_grup[i]) == kid for i in range(len(numery_grup))]
        sub = krzywe_df.iloc[:, maska]
        if sub.shape[1] == 0:
            continue
        sredni = sub.mean(axis=1); std = sub.std(axis=1).fillna(0)
        kolor = "#aaaaaa" if kid <= 0 else kolory_hex[(kid - 1) % 10]
        etykieta = "Szum" if kid <= 0 else f"Wzorzec Klastra {kid}"
        ax.fill_between(x, sredni - std, sredni + std, color=kolor, alpha=0.15)
        ax.plot(x, sredni, color=kolor, linewidth=1.8, label=etykieta)
    ax.set_xlabel("Oś X"); ax.set_ylabel("Sygnał (średnia ±1σ)")
    ax.legend(fontsize=7, framealpha=0.85)
    ax.grid(True, color="#e0e0e0", linewidth=0.5)
    ax.invert_xaxis()  # oś g-factor malejąca w prawo — konwencja EPR
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=dpi)
    plt.close(fig); buf.seek(0)
    return buf


def _png_dendrogram(dane, metoda, etykiety, dpi=300):
    try:
        fig, ax = plt.subplots(figsize=(8.0, 4.2))
        dendrogram(
            linkage(dane, method="ward" if "Warda" in metoda else "average"),
            labels=etykiety, leaf_rotation=90, ax=ax,
        )
        ax.set_ylabel("Odległość łączenia")
        fig.tight_layout()
        buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=dpi)
        plt.close(fig); buf.seek(0)
        return buf
    except Exception:
        return None


# =====================================================================
# RAMKI DANYCH X+Y DO EDYCJI W EXCELU
# Zamiast (lub obok) obrazka eksportujemy surowe serie: kolumna X oraz
# po jednej kolumnie Y na każdą krzywą / profil. Dzięki temu w Excelu
# można zmieniać kolory, typ linii, usuwać i dodawać serie.
# =====================================================================
def _df_krzywe_xy(x, krzywe_df, numery_grup, etykieciarz=None):
    """Surowe krzywe: kolumna X + po jednej kolumnie na krzywą.
    Nazwa kolumny zawiera nazwę krzywej i jej grupę, np. „y3 [Klaster A]”."""
    dane = {"X": np.asarray(x, dtype=float)}
    for i, col in enumerate(krzywe_df.columns):
        kid = int(numery_grup[i])
        etyk = etykieciarz(kid) if etykieciarz else (
            "Szum" if kid <= 0 else f"Klaster {kid}")
        dane[f"{col} [{etyk}]"] = krzywe_df.iloc[:, i].to_numpy()
    return pd.DataFrame(dane)


def _df_profile_xy(x, krzywe_df, numery_grup, etykieciarz=None):
    """Uśrednione profile: kolumna X + na każdą grupę kolumny średniej,
    dolnej i górnej granicy wstęgi (średnia ± 1 odch. std.)."""
    dane = {"X": np.asarray(x, dtype=float)}
    for kid in sorted(set(int(v) for v in numery_grup)):
        maska = [int(numery_grup[i]) == kid for i in range(len(numery_grup))]
        sub = krzywe_df.iloc[:, maska]
        if sub.shape[1] == 0:
            continue
        etyk = etykieciarz(kid) if etykieciarz else (
            "Szum" if kid <= 0 else f"Klaster {kid}")
        sredni = sub.mean(axis=1).to_numpy()
        std = sub.std(axis=1).fillna(0).to_numpy()
        dane[f"{etyk} — średnia"] = sredni
        dane[f"{etyk} — dolna (−1σ)"] = sredni - std
        dane[f"{etyk} — górna (+1σ)"] = sredni + std
    return pd.DataFrame(dane)


def _zapisz_df_z_szerokoscia(writer, df, nazwa_arkusza, szer=16):
    """Zapisuje DataFrame do arkusza i ustawia rozsądną szerokość kolumn."""
    df.to_excel(writer, index=False, sheet_name=nazwa_arkusza)
    ark = writer.sheets[nazwa_arkusza]
    from openpyxl.utils import get_column_letter
    for j in range(1, df.shape[1] + 1):
        ark.column_dimensions[get_column_letter(j)].width = szer
    return ark


def _df_klasyfikacja_xy(x_ref, macierz_wzorc_ujedn, klasy_ref,
                        dane_nowe_ujedn, nazwy_nowe, przypisania):
    """Dane wykresu nakładkowego klasyfikacji: kolumna X + uśrednione
    profile kategorii wzorcowych + każde nowe widmo jako osobna kolumna
    (z dopiskiem przypisanej kategorii). Gotowe do samodzielnej edycji."""
    dane = {"X (g-factor)": np.asarray(x_ref, dtype=float)}
    klasy_arr = np.asarray([str(k) for k in klasy_ref])
    for kat in sorted(set(klasy_arr)):
        maska = klasy_arr == kat
        dane[f"Wzorzec {kat}"] = macierz_wzorc_ujedn[maska].mean(axis=0)
    for j, nazwa in enumerate(nazwy_nowe):
        dane[f"{nazwa} → {przypisania[j]}"] = np.asarray(dane_nowe_ujedn[j])
    return pd.DataFrame(dane)


def _png_klasyfikacja(x_ref, macierz_wzorc_ujedn, klasy_ref,
                      dane_nowe_ujedn, nazwy_nowe, przypisania,
                      kolory_hex, dpi=300):
    try:
        klasy_arr = np.asarray([str(k) for k in klasy_ref])
        kategorie = sorted(set(klasy_arr))
        mapa = {kat: kolory_hex[i % 10] for i, kat in enumerate(kategorie)}
        fig, ax = plt.subplots(figsize=(7.2, 4.4))
        for kat in kategorie:
            maska = klasy_arr == kat
            ax.plot(x_ref, macierz_wzorc_ujedn[maska].mean(axis=0),
                    color=mapa[kat], linewidth=2.4, label=f"Wzorzec {kat}")
        for j, nazwa in enumerate(nazwy_nowe):
            kat_j = str(przypisania[j])
            ax.plot(x_ref, dane_nowe_ujedn[j], color=mapa.get(kat_j, "#aaaaaa"),
                    linewidth=1.1, linestyle="--", alpha=0.85)
        ax.set_xlabel("Oś X (g-factor)"); ax.set_ylabel("Sygnał")
        ax.legend(fontsize=7, framealpha=0.85)
        ax.grid(True, color="#e0e0e0", linewidth=0.5)
        ax.invert_xaxis()  # oś g-factor malejąca w prawo — konwencja EPR
        fig.tight_layout()
        buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=dpi)
        plt.close(fig); buf.seek(0)
        return buf
    except Exception:
        return None


def _hex(kolor):
    """Normalizuje kolor '#1f77b4' -> '1F77B4' na potrzeby openpyxl."""
    return str(kolor).lstrip("#").upper()


def _wykres_excel_klasyfikacja(ws, n_pkt, kolory_serii, style_serii,
                               tytul="Nowe widma na tle wzorców kategorii"):
    """Buduje natywny, edytowalny wykres Excela z danych już zapisanych w
    arkuszu `ws` (układ: kol.1 = X, kol.2..K = serie). Styl jak w aplikacji:
    profile wzorcowe — grube linie ciągłe; nowe widma — cienkie przerywane.

    kolory_serii: lista hex (bez #) w kolejności serii (kol. 2..K).
    style_serii:  lista 'solid' / 'dash' w tej samej kolejności.
    """
    n_serii = len(kolory_serii)
    chart = LineChart()
    chart.title = tytul
    chart.style = 2
    chart.height = 11        # cm
    chart.width = 20         # cm
    chart.x_axis.title = "Oś X (g-factor)"
    chart.y_axis.title = "Sygnał"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    # Oś g-factor malejąca w prawo — konwencja EPR
    chart.x_axis.scaling.orientation = "maxMin"

    dane = Reference(ws, min_col=2, max_col=1 + n_serii,
                     min_row=1, max_row=1 + n_pkt)
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_pkt)
    chart.add_data(dane, titles_from_data=True)
    chart.set_categories(cats)

    for idx, seria in enumerate(chart.series):
        dash = style_serii[idx] == "dash"
        lp = LineProperties(
            solidFill=kolory_serii[idx],
            w=12700 if dash else 28575,      # EMU: ~1 pt vs ~2.25 pt
        )
        if dash:
            lp.prstDash = "dash"
        seria.graphicalProperties = GraphicalProperties()
        seria.graphicalProperties.line = lp
        seria.smooth = False
    return chart


st.set_page_config(page_title="Analizator Krzywych Pro AI",
                   page_icon="⚛️", layout="wide")


# =====================================================================
# PRAWDZIWA SIEĆ SOM (SELF-ORGANIZING MAP) — mapa Kohonena
# z gaussowską funkcją sąsiedztwa i malejącym promieniem sigma
# =====================================================================

class SiecSOM:
    def __init__(self, x_size=5, y_size=5, input_dim=43,
                 lr=0.5, epochs=100, random_state=42):
        self.x_size = x_size
        self.y_size = y_size
        self.input_dim = input_dim
        self.lr = lr
        self.epochs = epochs
        self.random_state = random_state
        rng = np.random.default_rng(random_state)
        self.wagi = rng.random((x_size * y_size, input_dim))
        # Współrzędne neuronów na siatce 2D — potrzebne do sąsiedztwa
        gx, gy = np.meshgrid(np.arange(x_size), np.arange(y_size), indexing="ij")
        self.pozycje = np.column_stack([gx.ravel(), gy.ravel()]).astype(float)
        self.sigma0 = max(x_size, y_size) / 2.0

    def fit_predict_features(self, X):
        for epoch in range(self.epochs):
            frac = epoch / max(self.epochs - 1, 1)
            biezacy_lr = self.lr * (1.0 - frac)
            # Promień sąsiedztwa maleje od sigma0 do ~1 neuronu
            sigma = max(self.sigma0 * (1.0 - frac), 0.5)
            for sample in X:
                bmu_idx = np.argmin(np.linalg.norm(self.wagi - sample, axis=1))
                # Odległość każdego neuronu od BMU NA SIATCE (topologia!)
                dist_grid = np.linalg.norm(
                    self.pozycje - self.pozycje[bmu_idx], axis=1
                )
                h = np.exp(-(dist_grid ** 2) / (2.0 * sigma ** 2))
                # Aktualizacja wszystkich neuronów ważona sąsiedztwem
                self.wagi += biezacy_lr * h[:, None] * (sample - self.wagi)
        # Cechy = wagi BMU + drobny deterministyczny jitter, żeby uniknąć
        # identycznych wektorów (zerowe odległości psują linkage)
        aktywowane = np.empty((X.shape[0], self.input_dim))
        for i, sample in enumerate(X):
            bmu_idx = np.argmin(np.linalg.norm(self.wagi - sample, axis=1))
            aktywowane[i] = self.wagi[bmu_idx]
        rng = np.random.default_rng(self.random_state)
        aktywowane += rng.normal(0, 1e-9, aktywowane.shape)
        return aktywowane


# =====================================================================
# SŁOWNIKI OPISÓW
# =====================================================================

OPISY_METOD = {
    "Hierarchiczna Aglomeracyjna (metoda Warda)": "Buduje drzewo powiązań od dołu do góry na podstawie minimalizacji przyrostu wariancji wewnątrzklastrowej. Doskonale radzi sobie ze zwartymi grupami.",
    "Filtrowanie szumów (Rolling Mean) + Hierarchiczna (metoda Warda)": "Liniowa transformacja wygładzająca. Algorytm najpierw aplikuje okno kroczącej średniej (rolling mean), usuwając szum pomiarowy wysokiej częstotliwości z serii czasowej, a następnie grupuje klastry metodą Warda.",
    "PCA + Hierarchiczna (metoda Warda)": "Hybryda redukująca szum. Wyciąga kluczowe składowe sygnału (PCA), odrzucając drobne fluktuacje laboratoryjne, a następnie aplikuje kryterium Warda.",
    "UMAP + Hierarchiczna (metoda Warda)": "Potężna fuzja nieliniowa. UMAP makroskopowo zagęszcza i zbliża do siebie pokrewne profile krzywych w przestrzeni topologicznej, pozwalając metodzie Warda na bezbłędne wycięcie klastrów.",
    "SOM + Hierarchiczna (metoda Warda)": "Wykorzystuje topologiczną mapę Kohonena (SOM, z gaussowskim sąsiedztwem) do kompresji krzywych, a następnie buduje drzewo aglomeracyjne metodą Warda na bazie wag neuronów BMU.",
    "Spectral + Hierarchiczna (metoda Warda)": "Rzutuje krzywe do nieliniowej przestrzeni spektralnej (wektory własne Laplasjanu grafu podobieństwa), po czym aplikuje hierarchiczne grupowanie Warda na embeddingu.",
    "K-means": "Dzieli przestrzeń cech na tzw. obszary Voronoia. Algorytm dąży do minimalizacji wariancji wewnątrzklastrowej.",
    "UMAP + HDBSCAN (Hybryda Gęstościowa)": "Dwustopniowa hybryda nowej generacji. Najpierw rzutuje sygnał do przestrzeni topologicznej nieliniowej 2D (UMAP), a algorytm gęstościowy (HDBSCAN) wycina z nich grupy kształtów.",
    "Spectral + GMM (Hybryda Spektralno-Probabilistyczna)": "Najpierw wyznacza embedding spektralny (dekompozycja wartości własnych grafu pokrewieństwa), a następnie dopasowuje do niego elastyczne chmury probabilistyczne rozkładu normalnego (GMM).",
    "SOM + K-means (Hybryda sekwencyjna)": "Pierwszy etap wykorzystuje sieć neuronową Kohonena (SOM) do kompresji sygnału na siatkę topologiczną. Drugi etap uruchamia algorytm K-means na wagach neuronów.",
    "Klastrowanie Konsensusowe (Ensemble Voting)": "Metoda komitetowa. Uruchamia równolegle K-Means, GMM, Spectral, Ward i buduje macierz współwystępowania. Ostateczny podział jest fuzją decyzji wszystkich modeli.",
    "NMF (Nieujemna Faktoryzacja Macierzy)": "Rozkłada macierz danych na iloczyn dwóch macierzy o elementach wyłącznie nieujemnych.",
    "GMM (Probabilistyczna)": "Modele Mieszanin Gaussowskich. Próbuje dopasować elastyczne rozkłady normalne, dając miękkie przypisanie probabilistyczne.",
    "BGMM (Bayesowski GMM)": "Rozszerzenie GMM o probabilistyczną wersję Bayesowską z procesem Dirichleta. Automatycznie wygasza niepotrzebne klastry.",
    "Hierarchiczna Korelacyjna (metoda średnich)": "Podejście hierarchiczne, które mierzy stopień współliniowości wykresów za pomocą odległości korelacyjnej (1 - r Pearsona).",
    "HDBSCAN (Gęstościowa - Auto K)": "Zaawansowane klastrowanie gęstościowe oparte na teorii grafów. Szuka obszarów o wysokiej kondensacji punktów.",
    "Spectral Clustering": "Wykorzystuje wartości własne (widmo) macierzy podobieństwa danych do redukcji wymiarowości przed właściwym podziałem.",
    "K-Shape (Kształt fali)": "Wyspecjalizowany algorytm stworzony ściśle do analizy serii czasowych, wykorzystujący znormalizowaną korelację wzajemną.",
}

OPISY_PREPROCESSING = {
    "Standardowa": "Polega na klasycznej standaryzacji (Z-score). Sprowadza wszystkie punkty pomiarowe krzywych do wspólnej skali statystycznej.",
    "Analiza trendu": "Wyznacza różnice skończone (pochodne pierwszego rzędu) pomiędzy sąsiednimi punktami wzdłuż osi X.",
    "UMAP (Redukcja topologiczna)": "Uniform Manifold Approximation and Projection. Zaawansowana, nieliniowa redukcja wymiarowości.",
    "FeatureExtraction": "Głęboka transformacja inżynierska 3D: Max, Pozycja X, Średnia, Std, Skośność, Kurtoza, harmoniczne FFT oraz wskaźniki DWT Haar.",
    "MinMaxScaler": "Dokonuje liniowej transformacji danych, przesuwając i skalując wartości każdej krzywej do przedziału od 0 do 1.",
    "Filtrowanie szumów": "Wykorzystuje algorytm kroczącego okna średniej (rolling window). Skutecznie odcina fluktuacje wysokiej częstotliwości.",
    "Augmentacja sygnału": "Data Augmentation dla sieci neuronowych. Sztucznie rozbudowuje zbiór danych przez generowanie zaszumionych wariantów każdej krzywej. Dostępne techniki: Jitter (szum Gaussowski), Time Warping (deformacja osi czasu), Amplitude Scaling (losowe skalowanie amplitudy), Window Slicing (losowe przycięcie okna) oraz Permutation (przestawienie segmentów).",
}

OPISY_AUGMENTACJI = {
    "Jitter": "Dodaje do każdej krzywej losowy szum Gaussowski. Symuluje szum pomiarowy — najprostsza i najszybsza technika augmentacji.",
    "Time Warping": "Monotonicznie rozciąga i ściska oś czasu przez interpolację na odkształconej siatce punktów. Symuluje zmienną prędkość procesu.",
    "Amplitude Scaling": "Mnoży amplitudę każdej krzywej przez losowy współczynnik bliski 1.0. Symuluje zmienność wzmocnienia sygnału.",
    "Window Slicing": "Wycina losowy fragment krzywej i rozciąga go z powrotem do oryginalnej długości. Uczy model rozpoznawania lokalnych wzorców.",
    "Permutation": "Dzieli krzywą na segmenty i losowo je przestawia. Testuje odporność modelu na zmiany kolejności fragmentów sygnału.",
}


# =====================================================================
# FUNKCJE POMOCNICZE — DANE WEJŚCIOWE
# =====================================================================

# Wartości, które Excel wpisuje przy błędach formuł — traktujemy je jak puste.
_EXCEL_BLEDY = {
    "#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!",
    "#SPILL!", "#CALC!", "#GETTING_DATA", "#REF", "REF!",
}


class BladDanychWejsciowych(Exception):
    """Czytelny błąd wczytywania danych (pokazywany użytkownikowi po polsku)."""
    pass


def inteligentne_pobranie_tabeli(df_raw):
    # Zamień błędy Excela (#REF!, #DIV/0! itd.) na NaN — inaczej „zabetonują"
    # kolumnę jako tekstową i cała tabela wyjdzie pusta po konwersji na liczby.
    df_raw = df_raw.replace(list(_EXCEL_BLEDY), np.nan)
    # Obcięcie białych znaków w komórkach tekstowych i ponowna zamiana wariantów
    # błędów z ewentualnymi spacjami (np. " #REF! ").
    df_raw = df_raw.map(
        lambda v: (np.nan if isinstance(v, str) and v.strip().upper().rstrip("!")
                   in {b.rstrip("!") for b in _EXCEL_BLEDY} else v)
    )

    df_raw = df_raw.dropna(how="all", axis=0).dropna(how="all", axis=1)
    df_raw = df_raw.reset_index(drop=True)

    if df_raw.shape[0] < 2 or df_raw.shape[1] < 2:
        raise BladDanychWejsciowych(
            "Po odrzuceniu pustych komórek i błędów Excela (#REF!, #DIV/0! itp.) "
            "w arkuszu nie zostało dość danych. Sprawdź, czy plik nie zawiera "
            "zerwanych odwołań (#REF!) zamiast liczb — otwórz go w Excelu i "
            "zastąp formuły wartościami (Kopiuj → Wklej specjalnie → Wartości)."
        )

    indeks_startu = 0
    for idx, row in df_raw.iterrows():
        if row.notna().sum() > 1:
            if idx + 1 < len(df_raw):
                nastepny_wiersz = df_raw.iloc[idx + 1]
                ile_liczb = pd.to_numeric(nastepny_wiersz, errors="coerce").notna().sum()
                if ile_liczb > 1:
                    indeks_startu = idx
                    break
    naglowki = df_raw.iloc[indeks_startu]
    df_czysty = df_raw.iloc[indeks_startu + 1:].copy()
    df_czysty.columns = naglowki
    df_czysty = df_czysty.reset_index(drop=True)
    df_czysty = df_czysty.apply(pd.to_numeric, errors="coerce")
    df_czysty = df_czysty.dropna(how="all", axis=1)

    # Zabezpieczenie: jeśli po konwersji nie ma żadnej kolumny liczbowej,
    # dajemy zrozumiały komunikat zamiast kryptycznego IndexError.
    if df_czysty.shape[1] == 0:
        raise BladDanychWejsciowych(
            "Nie znaleziono ani jednej kolumny z danymi liczbowymi. Najczęstsza "
            "przyczyna to błędy formuł w pliku (#REF!, #DIV/0!, #VALUE! itp.), "
            "które zastąpiły liczby. Otwórz plik w Excelu, napraw lub usuń "
            "zerwane odwołania i zapisz jako wartości, a następnie wgraj ponownie."
        )

    df_czysty = df_czysty.dropna(subset=[df_czysty.columns[0]])

    if df_czysty.shape[0] == 0:
        raise BladDanychWejsciowych(
            "Pierwsza kolumna (oś X) nie zawiera żadnych wartości liczbowych po "
            "oczyszczeniu. Sprawdź, czy kolumna g-factor / osi X nie składa się "
            "z błędów formuł lub tekstu."
        )

    return df_czysty.reset_index(drop=True)


@st.cache_data(show_spinner="Pobieram dane z Google Sheets...", max_entries=3, ttl=1800)
def pobierz_google_sheets(url_base):
    """Jeden request HTTP dla wszystkich arkuszy naraz."""
    return pd.read_excel(f"{url_base}/export?format=xlsx", sheet_name=None, header=None)


# =====================================================================
# AUGMENTACJA SYGNAŁU (poprawiony Time Warping)
# =====================================================================

def augmentuj_sygnal(krzywe_df, technika, sila, n_kopii, random_state=42):
    """
    Generuje n_kopii augmentowanych wariantów każdej krzywej i zwraca
    rozszerzony DataFrame wraz z etykietami źródłowymi.
    sila: float 0.0–1.0 — intensywność przekształcenia.
    Kolejność kolumn: NAJPIERW wszystkie oryginały, potem kopie.
    """
    rng = np.random.default_rng(random_state)
    wyniki = {}
    etykiety_zrodlowe = {}

    for col in krzywe_df.columns:
        wyniki[str(col)] = krzywe_df[col].values.copy()
        etykiety_zrodlowe[str(col)] = str(col)

    n_punktow = len(krzywe_df)
    t_orig = np.linspace(0, 1, n_punktow)

    for col in krzywe_df.columns:
        syg = krzywe_df[col].values.astype(float)

        for k in range(1, n_kopii + 1):
            nazwa_aug = f"{col}_aug{k}"

            if technika == "Jitter":
                szum = rng.normal(0, sila * np.std(syg), size=n_punktow)
                aug = syg + szum

            elif technika == "Time Warping":
                # Monotoniczne odkształcenie osi czasu:
                # węzły równomierne -> losowo przesunięte -> wymuszona
                # monotoniczność -> interpolacja mapy deformacji.
                n_wezlow = 6
                t_wezly = np.linspace(0, 1, n_wezlow)
                przesuniecia = rng.uniform(-sila * 0.3, sila * 0.3, n_wezlow)
                przesuniecia[0] = 0.0
                przesuniecia[-1] = 0.0
                t_zdeform = np.clip(t_wezly + przesuniecia, 0, 1)
                t_zdeform = np.maximum.accumulate(t_zdeform)  # monotonia
                t_zdeform[-1] = 1.0
                mapa_czasu = np.interp(t_orig, t_wezly, t_zdeform)
                aug = np.interp(mapa_czasu, t_orig, syg)

            elif technika == "Amplitude Scaling":
                wspolczynnik = rng.uniform(1.0 - sila * 0.5, 1.0 + sila * 0.5)
                aug = syg * wspolczynnik

            elif technika == "Window Slicing":
                min_dlugosc = max(int(n_punktow * (1.0 - sila * 0.4)), 3)
                if min_dlugosc >= n_punktow:
                    aug = syg.copy()
                else:
                    dlugosc_okna = int(rng.integers(min_dlugosc, n_punktow))
                    start = int(rng.integers(0, n_punktow - dlugosc_okna + 1))
                    wycinek = syg[start: start + dlugosc_okna]
                    aug = np.interp(
                        t_orig, np.linspace(0, 1, dlugosc_okna), wycinek
                    )

            elif technika == "Permutation":
                n_segmentow = max(2, int(2 + sila * 8))
                granice = np.array_split(np.arange(n_punktow), n_segmentow)
                kolejnosc = rng.permutation(len(granice))
                aug = np.concatenate([syg[granice[i]] for i in kolejnosc])

            else:
                aug = syg.copy()

            wyniki[nazwa_aug] = aug
            etykiety_zrodlowe[nazwa_aug] = str(col)

    df_aug = pd.DataFrame(wyniki, index=krzywe_df.index)
    return df_aug, etykiety_zrodlowe


# =====================================================================
# SILNIK KLASTROWANIA
# UWAGA: df_sygnaly_raw jest teraz częścią klucza cache'a (bez "_"),
# więc zmiana surowych danych zawsze unieważnia cache.
# =====================================================================

def _spectral_embedding(dane, n_komponentow, random_state=42):
    """Wspólny embedding spektralny dla hybryd Spectral+X."""
    n_comp = max(2, min(n_komponentow, dane.shape[0] - 2))
    emb = SpectralEmbedding(
        n_components=n_comp, affinity="rbf", random_state=random_state
    )
    return emb.fit_transform(dane)


@st.cache_data(show_spinner=False, max_entries=60, ttl=3600)
def uruchom_silnik_klastrowania(nazwa_metody, dane, k_grup, min_hdbscan=3,
                                df_sygnaly_raw=None, seed=None):
    # seed=None -> zachowanie jak dotąd (deterministyczne, KONFIG). Podanie
    # seed pozwala rankingowi uruchomić tę samą metodę na różnych ziarnach
    # i policzyć stabilność (średnia ± odchylenie).
    rs = KONFIG["RANDOM_STATE"] if seed is None else int(seed)

    if nazwa_metody == "K-means":
        return KMeans(n_clusters=k_grup, random_state=rs, n_init=5).fit_predict(dane) + 1

    elif "Filtrowanie szumów (Rolling Mean) + Hierarchiczna" in nazwa_metody:
        if df_sygnaly_raw is not None:
            wygladzane = df_sygnaly_raw.rolling(
                window=KONFIG["ROLLING_WINDOW"], center=True, min_periods=1
            ).mean().T
            dane_ward = StandardScaler().fit_transform(wygladzane)
        else:
            dane_ward = dane
        return fcluster(linkage(dane_ward, method="ward"), t=k_grup, criterion="maxclust")

    elif "PCA + Hierarchiczna" in nazwa_metody:
        n_comp = min(KONFIG["PCA_KOMPONENTY"], dane.shape[1], dane.shape[0])
        komponenty_pca = PCA(n_components=n_comp, random_state=rs).fit_transform(dane)
        return fcluster(linkage(komponenty_pca, method="ward"), t=k_grup, criterion="maxclust")

    elif "UMAP + Hierarchiczna" in nazwa_metody and umap_dostepne:
        przestrzen_2d = umap.UMAP(
            n_neighbors=KONFIG["UMAP_N_NEIGHBORS"],
            min_dist=KONFIG["UMAP_MIN_DIST"], random_state=rs
        ).fit_transform(dane)
        return fcluster(linkage(przestrzen_2d, method="ward"), t=k_grup, criterion="maxclust")

    elif "SOM + Hierarchiczna" in nazwa_metody:
        model_som = SiecSOM(
            x_size=KONFIG["SOM_X"], y_size=KONFIG["SOM_Y"],
            input_dim=dane.shape[1], lr=KONFIG["SOM_LR"],
            epochs=KONFIG["SOM_EPOKI"], random_state=rs
        )
        cechy_som = model_som.fit_predict_features(dane)
        return fcluster(linkage(cechy_som, method="ward"), t=k_grup, criterion="maxclust")

    elif "Spectral + Hierarchiczna" in nazwa_metody:
        # POPRAWKA: Ward na embeddingu spektralnym, nie na macierzy afiniczności
        emb = _spectral_embedding(dane, k_grup, rs)
        return fcluster(linkage(emb, method="ward"), t=k_grup, criterion="maxclust")

    elif "UMAP + HDBSCAN" in nazwa_metody and umap_dostepne:
        baza_projekcji = StandardScaler().fit_transform(dane)
        przestrzen_2d = umap.UMAP(
            n_neighbors=KONFIG["UMAP_N_NEIGHBORS"],
            min_dist=KONFIG["UMAP_MIN_DIST"], random_state=rs
        ).fit_transform(baza_projekcji)
        raw_labels = HDBSCAN(min_cluster_size=min_hdbscan, min_samples=1).fit_predict(przestrzen_2d)
        return np.array([n + 1 if n >= 0 else 0 for n in raw_labels])

    elif "Spectral + GMM" in nazwa_metody:
        # POPRAWKA: GMM na embeddingu spektralnym (wcześniej Spectral
        # w ogóle nie był używany)
        emb = _spectral_embedding(dane, k_grup, rs)
        gmm = GaussianMixture(n_components=k_grup, random_state=rs, n_init=2)
        return gmm.fit_predict(emb) + 1

    elif "SOM + K-means" in nazwa_metody:
        model_som = SiecSOM(
            x_size=KONFIG["SOM_X"], y_size=KONFIG["SOM_Y"],
            input_dim=dane.shape[1], lr=KONFIG["SOM_LR"],
            epochs=KONFIG["SOM_EPOKI"], random_state=rs
        )
        cechy_som = model_som.fit_predict_features(dane)
        return KMeans(n_clusters=k_grup, random_state=rs, n_init=5).fit_predict(cechy_som) + 1

    elif "Konsensusowe" in nazwa_metody:
        N = dane.shape[0]
        matrix = np.zeros((N, N))
        p1 = KMeans(n_clusters=k_grup, random_state=rs, n_init=2).fit_predict(dane)
        p2 = GaussianMixture(n_components=k_grup, random_state=rs, n_init=1).fit_predict(dane)
        p3 = SpectralClustering(n_clusters=k_grup, random_state=rs,
                                assign_labels="discretize").fit_predict(dane)
        p4 = fcluster(linkage(dane, method="ward"), t=k_grup, criterion="maxclust") - 1
        # POPRAWKA WYDAJNOŚCI: wektoryzacja zamiast potrójnej pętli
        for p in [p1, p2, p3, p4]:
            p = np.asarray(p)
            matrix += (p[:, None] == p[None, :]).astype(float)
        # POPRAWKA: linkage wymaga skondensowanej macierzy odległości —
        # kwadratowa macierz byłaby zinterpretowana jak zwykłe obserwacje
        dyssymilarnosc = 1.0 - (matrix / 4.0)
        np.fill_diagonal(dyssymilarnosc, 0.0)
        link = linkage(squareform(dyssymilarnosc, checks=False), method="average")
        return fcluster(link, t=k_grup, criterion="maxclust")

    elif "NMF" in nazwa_metody:
        dane_nmf = MinMaxScaler().fit_transform(dane) if (dane < 0).any() else dane
        W = NMF(n_components=k_grup, init="nndsvd", random_state=rs,
                max_iter=200).fit_transform(dane_nmf)
        return np.argmax(W, axis=1) + 1

    elif nazwa_metody == "GMM (Probabilistyczna)":
        return GaussianMixture(n_components=k_grup, random_state=rs,
                               n_init=2).fit_predict(dane) + 1

    elif "BGMM" in nazwa_metody:
        return BayesianGaussianMixture(
            n_components=k_grup, covariance_type="diag",
            weight_concentration_prior=1e-3, random_state=rs, n_init=2
        ).fit_predict(dane) + 1

    elif "metoda Warda" in nazwa_metody:
        return fcluster(linkage(dane, method="ward"), t=k_grup, criterion="maxclust")

    elif "Korelacyjna" in nazwa_metody:
        return fcluster(linkage(dane, method="average", metric="correlation"),
                        t=k_grup, criterion="maxclust")

    elif nazwa_metody == "HDBSCAN (Gęstościowa - Auto K)":
        raw = HDBSCAN(min_cluster_size=min_hdbscan, min_samples=1).fit_predict(dane)
        return np.array([n + 1 if n >= 0 else 0 for n in raw])

    elif "Spectral Clustering" in nazwa_metody:
        return SpectralClustering(n_clusters=k_grup, random_state=rs,
                                  assign_labels="discretize").fit_predict(dane) + 1

    elif "K-Shape" in nazwa_metody and tslearn_dostepne:
        return KShape(n_clusters=k_grup, random_state=rs).fit_predict(
            to_time_series_dataset(dane)) + 1

    else:
        return fcluster(linkage(dane, method="ward"), t=k_grup, criterion="maxclust")


# =====================================================================
# CACHE'OWANE OBLICZENIA POMOCNICZE
# =====================================================================

@st.cache_data(show_spinner=False, max_entries=10, ttl=3600)
def oblicz_metryki_k(dane, k_min, k_max):
    inercje, silhouettes, db_scores, calinski = [], [], [], []
    zakres_k = list(range(k_min, k_max + 1))
    for k in zakres_k:
        km = KMeans(n_clusters=k, random_state=KONFIG["RANDOM_STATE"], n_init=5).fit(dane)
        labels = km.labels_
        inercje.append(km.inertia_)
        silhouettes.append(silhouette_score(dane, labels))
        db_scores.append(davies_bouldin_score(dane, labels))
        calinski.append(calinski_harabasz_score(dane, labels))
    return zakres_k, inercje, silhouettes, db_scores, calinski


@st.cache_data(show_spinner="Analiza Leave-One-Out w toku... (N pełnych klasteryzacji)", max_entries=3, ttl=1800)
def oblicz_leave_one_out(metoda, dane_loo, etykiety_gt, krzywe_raw,
                         nazwy, liczba_grup, ari_bazowe):
    """LOO cache'owane i uruchamiane wyłącznie na żądanie (przycisk)."""
    wyniki_loo = []
    N = dane_loo.shape[0]
    for odrzucona_idx in range(N):
        maska = np.ones(N, dtype=bool)
        maska[odrzucona_idx] = False
        dane_sub = dane_loo[maska]
        etykiety_sub = [etykiety_gt[i] for i in range(N) if maska[i]]
        if "Filtrowanie szumów (Rolling Mean) + Hierarchiczna" in metoda:
            krzywe_sub = krzywe_raw.iloc[:, maska]
        else:
            krzywe_sub = krzywe_raw
        pred_sub = uruchom_silnik_klastrowania(
            metoda, dane_sub, liczba_grup, liczba_grup, df_sygnaly_raw=krzywe_sub
        )
        sub_ari = adjusted_rand_score(etykiety_sub, pred_sub) * 100
        sub_nmi = normalized_mutual_info_score(etykiety_sub, pred_sub) * 100
        wyniki_loo.append({
            "Odrzucona Krzywa": str(nazwy[odrzucona_idx]),
            "ARI bez krzywej (%)": round(sub_ari, 2),
            "NMI bez krzywej (%)": round(sub_nmi, 2),
            "Wpływ na ARI": round(sub_ari - ari_bazowe, 2),
        })
    return wyniki_loo


@st.cache_data(show_spinner="Obliczam ranking (wiele ziaren)...", max_entries=3, ttl=1800)
def oblicz_ranking(dane, etykiety, krzywe_raw, metody_tuple, k,
                   indeksy_oryginalow, n_ziaren=5):
    """
    Ranking wszystkich metod z oceną STABILNOŚCI. Każda metoda jest
    uruchamiana na kilku ziarnach; raportujemy średnią ± odchylenie.

    OCHRONA PAMIĘCI (kluczowe na Streamlit Cloud, ~1 GB):
    - Metody DETERMINISTYCZNE (Ward, PCA+Ward, korelacyjna, NMF) liczone są
      tylko RAZ — dają σ=0 z definicji, więc N przebiegów to marnotrawstwo.
    - Metody CIĘŻKIE (UMAP, Spectral, SOM) dostają ograniczoną liczbę ziaren
      (max 3), bo to one zjadają pamięć — kilkanaście ich instancji naraz
      wywoływało segmentation fault (crash OOM w numbie/UMAP).
    - Lekkie losowe (K-means, GMM, BGMM) dostają pełne N ziaren.
    - Pamięć jest zwalniana (gc) po każdej metodzie.

    Przy augmentacji predykcje są przycinane do oryginalnych krzywych przed
    liczeniem ARI/NMI (POPRAWKA #6). etykiety=None -> tylko silhouette.
    """
    rekordy = []
    bledy = []
    idx_oryg = list(indeksy_oryginalow) if indeksy_oryginalow is not None else None
    n_ziaren = max(1, int(n_ziaren))

    # Klasyfikacja metody -> ile ziaren realnie potrzebuje.
    _DETERMINISTYCZNE = ("metoda Warda", "PCA + Hierarchiczna", "Korelacyjna",
                         "NMF", "Filtrowanie szumów (Rolling Mean)")
    _CIEZKIE = ("UMAP", "Spectral", "SOM", "Konsensusowe", "K-Shape")

    def _ile_ziaren(nazwa):
        if any(d in nazwa for d in _DETERMINISTYCZNE):
            return 1                      # deterministyczna: σ=0, jeden przebieg
        if any(c in nazwa for c in _CIEZKIE):
            return min(n_ziaren, 3)       # ciężka: limit dla ochrony pamięci
        return n_ziaren                   # lekka losowa: pełne N

    for m_nazwa in metody_tuple:
        ziarna = [KONFIG["RANDOM_STATE"] + 7 * s
                  for s in range(_ile_ziaren(m_nazwa))]
        ari_lista, nmi_lista, sil_lista = [], [], []
        ostatni_blad = None
        n_nieudanych = 0

        for sd in ziarna:
            try:
                pred = uruchom_silnik_klastrowania(
                    m_nazwa, dane, k, k, df_sygnaly_raw=krzywe_raw, seed=sd
                )
                pred = np.asarray(pred)
                unikalne = np.unique(pred[pred > 0]) if 0 in pred else np.unique(pred)
                if len(unikalne) < 2:
                    n_nieudanych += 1
                    ostatni_blad = "mniej niż 2 klastry"
                    continue

                # Silhouette na pełnym zbiorze (z pominięciem szumu HDBSCAN)
                maska_ns = pred > 0
                if maska_ns.sum() >= 2 and len(np.unique(pred[maska_ns])) >= 2:
                    m_sil = silhouette_score(dane[maska_ns], pred[maska_ns]) * 100
                else:
                    m_sil = silhouette_score(dane, pred) * 100
                sil_lista.append(m_sil)

                if etykiety is not None:
                    pred_gt = pred[idx_oryg] if idx_oryg is not None else pred
                    ari_lista.append(adjusted_rand_score(list(etykiety), pred_gt) * 100)
                    nmi_lista.append(normalized_mutual_info_score(list(etykiety), pred_gt) * 100)
            except Exception as e:
                n_nieudanych += 1
                ostatni_blad = str(e)

        gc.collect()  # zwolnij pamięć po każdej metodzie (ochrona przed OOM)

        if not sil_lista:
            bledy.append(f"{m_nazwa}: {ostatni_blad or 'brak wyników'}")
            continue

        rekord = {"Algorytm AI": m_nazwa}

        def _sr_od(lista):
            arr = np.asarray(lista, dtype=float)
            return round(float(arr.mean()), 2), round(float(arr.std()), 2)

        sil_sr, sil_od = _sr_od(sil_lista)
        rekord["Silhouette (%)"] = sil_sr
        rekord["Silhouette σ"] = sil_od

        if etykiety is not None and ari_lista:
            ari_sr, ari_od = _sr_od(ari_lista)
            nmi_sr, nmi_od = _sr_od(nmi_lista)
            rekord["ARI (%)"] = ari_sr
            rekord["ARI σ"] = ari_od
            rekord["NMI (%)"] = nmi_sr
            rekord["NMI σ"] = nmi_od
            rekord["Średnia (%)"] = round((ari_sr + nmi_sr + sil_sr) / 3, 2)
            # Łączna niestabilność = pierwiastek ze średniej wariancji metryk
            rekord["Rozrzut σ"] = round(
                float(np.sqrt(np.mean([ari_od**2, nmi_od**2, sil_od**2]))), 2)
        else:
            rekord["Średnia (%)"] = sil_sr
            rekord["Rozrzut σ"] = sil_od

        rekord["Udane pow."] = f"{len(sil_lista)}/{len(ziarna)}"
        rekordy.append(rekord)

    return rekordy, bledy


# =====================================================================
# GŁÓWNY EKRAN APLIKACJI
# =====================================================================

st.title("📊 Interaktywny Analizator Krzywych AI Pro")
st.write("### Ustawienia analizy")

# --- Panel zarządzania pamięcią (zawsze dostępny) ---
with st.sidebar:
    with st.expander("⚙️ Pamięć i wydajność", expanded=False):
        st.caption(
            "Aplikacja buforuje kosztowne obliczenia (klastrowanie, ranking, "
            "Leave-One-Out), aby przyspieszyć pracę. Bufor ma ograniczony "
            "rozmiar i sam zwalnia stare wyniki. Jeśli aplikacja zwalnia lub "
            "zużywa dużo pamięci, możesz wyczyścić bufor ręcznie."
        )
        if st.button("🧹 Wyczyść pamięć podręczną (cache)",
                     width='stretch'):
            st.cache_data.clear()
            gc.collect()
            st.success("Pamięć podręczna wyczyszczona.")

typ_zrodla = st.radio(
    "Wybierz źródło danych:",
    ["Plik Excel (.xlsx)", "Link do Google Sheets"],
    horizontal=True,
)

df = None
df_expert_raw = None
file_id = "default"
nazwa_pliku_wzorcowego = None  # do opisowych nazw eksportów
nazwa_pliku_nowych = None      # nazwa pliku z nowymi widmami (klasyfikacja)

if typ_zrodla == "Plik Excel (.xlsx)":
    uploaded_file = st.file_uploader("Wgraj plik Excel", type=["xlsx"])
    if uploaded_file is not None:
        nazwa_pliku_wzorcowego = uploaded_file.name
        try:
            df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
            df = inteligentne_pobranie_tabeli(df_raw)
        except BladDanychWejsciowych as e:
            st.error(f"⚠️ Nie udało się wczytać danych z pliku.\n\n{e}")
            st.stop()
        except Exception as e:
            st.error(
                "⚠️ Nie udało się odczytać pliku Excel. Sprawdź, czy to "
                f"prawidłowy plik .xlsx z danymi.\n\nSzczegóły: {e}"
            )
            st.stop()
        file_id = f"local_{len(df_raw)}_{df_raw.shape[1]}_{uploaded_file.size}"
        try:
            excel_file = pd.ExcelFile(uploaded_file)
            if "Ground Truth" in excel_file.sheet_names:
                df_expert_raw = pd.read_excel(uploaded_file, sheet_name="Ground Truth")
            elif len(excel_file.sheet_names) > 1:
                df_expert_raw = pd.read_excel(uploaded_file, sheet_name=1)
        except Exception:
            df_expert_raw = None
else:
    link_sheets = st.text_input(
        "Wklej link do Google Sheets:",
        placeholder="https://docs.google.com/spreadsheets/d/...",
    )
    if link_sheets and "docs.google.com/spreadsheets" in link_sheets:
        nazwa_pliku_wzorcowego = "GoogleSheets"
        try:
            url_base = link_sheets.split("/edit")[0]
            # POPRAWKA: jeden request zamiast dwóch
            sheets_dict = pobierz_google_sheets(url_base)
            nazwy_arkuszy = list(sheets_dict.keys())
            df = inteligentne_pobranie_tabeli(sheets_dict[nazwy_arkuszy[0]])
            file_id = f"cloud_{link_sheets[-15:]}"
            if "Ground Truth" in sheets_dict:
                gt_raw = sheets_dict["Ground Truth"]
                gt_raw.columns = gt_raw.iloc[0]
                df_expert_raw = gt_raw.iloc[1:].reset_index(drop=True)
            elif len(nazwy_arkuszy) > 1:
                gt_raw = sheets_dict[nazwy_arkuszy[1]]
                gt_raw.columns = gt_raw.iloc[0]
                df_expert_raw = gt_raw.iloc[1:].reset_index(drop=True)
        except BladDanychWejsciowych as e:
            st.error(f"⚠️ Nie udało się wczytać danych z arkusza.\n\n{e}")
            st.stop()
        except Exception:
            st.error("Nie udało się pobrać danych ze struktur Google Sheets.")

if df is None:
    st.info("Aby rozpocząć, wgraj plik z dysku lub wklej link do Google Sheets powyżej.")
    st.stop()

try:
    x = df.iloc[:, 0]
    krzywe = df.iloc[:, 1:]
    nazwy_krzywych = krzywe.columns.tolist()
    n_krzywych = len(nazwy_krzywych)

    if n_krzywych < 3:
        st.error("Za mało krzywych do analizy (wymagane minimum 3).")
        st.stop()

    # Podgląd wczytanych danych — heurystyka nagłówka przestaje być czarną skrzynką
    with st.expander("🗂️ Podgląd wczytanych danych", expanded=False):
        st.caption(
            f"Wykryto **{n_krzywych}** krzywych po **{len(df)}** punktów. "
            f"Pierwsza kolumna (`{df.columns[0]}`) traktowana jako oś X."
        )
        st.dataframe(df.head(8), width='stretch')

    # -----------------------------------------------------------------
    # LISTY METOD I PREPROCESSINGU
    # -----------------------------------------------------------------
    lista_metod = [
        "Hierarchiczna Aglomeracyjna (metoda Warda)",
        "Filtrowanie szumów (Rolling Mean) + Hierarchiczna (metoda Warda)",
        "PCA + Hierarchiczna (metoda Warda)",
        "SOM + Hierarchiczna (metoda Warda)",
        "Spectral + Hierarchiczna (metoda Warda)",
        "K-means",
        "Spectral + GMM (Hybryda Spektralno-Probabilistyczna)",
        "SOM + K-means (Hybryda sekwencyjna)",
        "Klastrowanie Konsensusowe (Ensemble Voting)",
        "NMF (Nieujemna Faktoryzacja Macierzy)",
        "GMM (Probabilistyczna)",
        "BGMM (Bayesowski GMM)",
        "Hierarchiczna Korelacyjna (metoda średnich)",
        "HDBSCAN (Gęstościowa - Auto K)",
        "Spectral Clustering",
    ]
    if umap_dostepne:
        lista_metod.insert(3, "UMAP + Hierarchiczna (metoda Warda)")
        lista_metod.insert(7, "UMAP + HDBSCAN (Hybryda Gęstościowa)")
    if tslearn_dostepne:
        lista_metod.append("K-Shape (Kształt fali)")

    lista_preprocessingow = ["Standardowa", "Analiza trendu"]
    if umap_dostepne:
        lista_preprocessingow.append("UMAP (Redukcja topologiczna)")
    lista_preprocessingow.extend(
        ["FeatureExtraction", "MinMaxScaler", "Filtrowanie szumów", "Augmentacja sygnału"]
    )

    if ("wybrana_metoda" not in st.session_state
            or st.session_state.wybrana_metoda not in lista_metod):
        st.session_state.wybrana_metoda = lista_metod[0]

    # Zakres K ograniczony realną liczbą krzywych (POPRAWKA #14)
    max_k = min(KONFIG["K_MAX_SUGESTIA"], n_krzywych - 1)

    col_param1, col_param2, col_param3 = st.columns(3)
    with col_param1:
        metoda = st.selectbox("Wybierz metodę główną:", lista_metod, key="wybrana_metoda")
    with col_param2:
        if ("K-Shape" in metoda or "UMAP + HDBSCAN" in metoda
                or "UMAP + Hierarchiczna" in metoda):
            optymalizacja = "Standardowa"
            st.caption("Ta metoda używa standardowego przygotowania danych.")
        else:
            optymalizacja = st.selectbox(
                "Wybierz wstępne przygotowanie danych:", lista_preprocessingow
            )
    with col_param3:
        if "HDBSCAN" in metoda:
            slider_label = "Minimalna wielkość grupy (HDBSCAN):"
        elif "BGMM" in metoda:
            slider_label = "Maksymalna liczba grup (BGMM):"
        else:
            slider_label = "Liczba grup (K):"
        liczba_grup = st.slider(slider_label, min_value=2, max_value=max_k,
                                value=min(5, max_k))

    st.write("---")

    # -----------------------------------------------------------------
    # PANEL AUGMENTACJI
    # -----------------------------------------------------------------
    aug_technika = "Jitter"
    aug_sila = 0.1
    aug_kopie = 2

    if optymalizacja == "Augmentacja sygnału":
        with st.expander("⚙️ Ustawienia Augmentacji Sygnału", expanded=False):
            col_aug1, col_aug2, col_aug3 = st.columns(3)
            with col_aug1:
                aug_technika = st.selectbox(
                    "Technika augmentacji:",
                    ["Jitter", "Time Warping", "Amplitude Scaling",
                     "Window Slicing", "Permutation"],
                    help="Wybierz metodę przekształcania krzywych",
                )
                st.caption(OPISY_AUGMENTACJI.get(aug_technika, ""))
            with col_aug2:
                aug_sila = st.slider(
                    "Siła augmentacji:",
                    min_value=0.01, max_value=1.0, value=0.1, step=0.01,
                    help="Im wyższa wartość, tym większe zniekształcenie sygnału",
                )
            with col_aug3:
                aug_kopie = st.slider(
                    "Liczba kopii na krzywą:",
                    min_value=1, max_value=10, value=2,
                    help="Ile augmentowanych wariantów wygenerować dla każdej krzywej",
                )
            st.info(
                f"Zbiór zostanie rozszerzony z **{n_krzywych}** do "
                f"**{n_krzywych * (1 + aug_kopie)}** krzywych "
                f"({aug_kopie} kopii × {n_krzywych} oryginałów + oryginały). "
                f"Klasteryzacja działa na pełnym zbiorze, ARI/NMI liczone tylko dla oryginałów."
            )

    # =================================================================
    # GROUND TRUTH — uczciwa obsługa (POPRAWKA #8)
    # ARI/NMI liczone tylko gdy istnieje realny podział ekspercki.
    # =================================================================
    expert_mapping = {}
    if df_expert_raw is not None and len(df_expert_raw) > 0:
        try:
            df_expert_raw.columns = [str(c).strip().lower() for c in df_expert_raw.columns]
            col_k = df_expert_raw.columns[0]
            col_g = df_expert_raw.columns[1]
            for _, row in df_expert_raw.iterrows():
                k_str = str(row[col_k]).strip().lower()
                v_str = str(row[col_g]).strip()
                if k_str and k_str != "nan" and v_str and v_str != "nan":
                    expert_mapping[k_str] = v_str
        except Exception:
            expert_mapping = {}

    # Domyślny (historyczny) podział y1–y43 — używany TYLKO gdy nazwy pasują
    sztywny_podzial_eksperta = {}
    for i in range(1, 44):
        if i <= 17:
            sztywny_podzial_eksperta[f"y{i}"] = "a"
        elif i <= 21:
            sztywny_podzial_eksperta[f"y{i}"] = "b"
        elif i <= 35:
            sztywny_podzial_eksperta[f"y{i}"] = "c"
        else:
            sztywny_podzial_eksperta[f"y{i}"] = "e"

    def znajdz_gt(nazwa, mapa):
        n_clean = str(nazwa).strip().lower()
        n_alt = f"y{n_clean}" if (not n_clean.startswith("y") and n_clean.isdigit()) else n_clean
        if n_clean in mapa:
            return mapa[n_clean]
        if n_alt in mapa:
            return mapa[n_alt]
        return None

    expert_list = []
    zrodlo_gt = None  # "plik" | "domyslny" | None

    if expert_mapping:
        trafienia = [znajdz_gt(n, expert_mapping) for n in nazwy_krzywych]
        if all(t is not None for t in trafienia):
            expert_list = trafienia
            zrodlo_gt = "plik"
    if zrodlo_gt is None:
        trafienia = [znajdz_gt(n, sztywny_podzial_eksperta) for n in nazwy_krzywych]
        if all(t is not None for t in trafienia):
            expert_list = trafienia
            zrodlo_gt = "domyslny"

    gt_dostepny = zrodlo_gt is not None

    # -----------------------------------------------------------------
    # SIDEBAR — edytor Grup Wzorcowych (tylko gdy GT dostępny)
    # -----------------------------------------------------------------
    etykiety_eksperta = None
    if gt_dostepny:
        df_current_gt = pd.DataFrame({
            "Krzywa": [str(n) for n in nazwy_krzywych],
            "Grupa Eksperta": expert_list,
        })

        if ("last_file_id" not in st.session_state
                or st.session_state.last_file_id != file_id):
            st.session_state.last_file_id = file_id
            st.session_state["tabela_editor_state"] = df_current_gt

        with st.sidebar:
            st.markdown("### 📋 Grupy Wzorcowe")
            if zrodlo_gt == "domyslny":
                st.warning(
                    "⚠️ Brak arkusza *Ground Truth* — użyto **domyślnego** "
                    "podziału y1–y43. Zweryfikuj przed interpretacją ARI/NMI!"
                )
            st.caption("Zmiana grupy natychmiast przelicza ARI, NMI i anomalie.")
            edited_gt = st.data_editor(
                st.session_state["tabela_editor_state"],
                hide_index=True,
                width='stretch',
                disabled=["Krzywa"],
                key=f"sidebar_editor_{file_id}",
                column_config={
                    "Krzywa": st.column_config.TextColumn("Krzywa", disabled=True),
                    "Grupa Eksperta": st.column_config.TextColumn(
                        "Grupa", help="Wpisz nazwę grupy (a, b, c...)", max_chars=20
                    ),
                },
            )
            st.session_state["tabela_editor_state"] = edited_gt

        etykiety_eksperta = edited_gt["Grupa Eksperta"].astype(str).tolist()
    else:
        with st.sidebar:
            st.markdown("### 📋 Grupy Wzorcowe")
            st.info(
                "Brak podziału eksperckiego (arkusz *Ground Truth* z kolumnami "
                "Krzywa | Grupa). Metryki ARI/NMI są wyłączone — dostępne "
                "pozostają metryki wewnętrzne (Silhouette, Davies-Bouldin...)."
            )

    # -----------------------------------------------------------------
    # Stylizacja sidebara: niebieski pasek „Grupy Wzorcowe" na zwiniętym
    # przycisku + kursory wskazujące na kontrolkach. Czysty CSS (bez JS).
    # -----------------------------------------------------------------
    st.markdown("""
    <style>
    [data-testid="collapsedControl"] {
        background: #1f77b4 !important;
        border-radius: 0 8px 8px 0 !important;
        width: 34px !important;
        height: auto !important;
        min-height: 80px !important;
        cursor: pointer !important;
        overflow: visible !important;
    }
    [data-testid="collapsedControl"] svg { display: none !important; }
    [data-testid="collapsedControl"]::after {
        content: 'Grupy Wzorcowe';
        color: white;
        font-size: 11px;
        font-weight: bold;
        writing-mode: vertical-rl;
        letter-spacing: 1px;
        padding: 10px 4px;
        display: block;
        text-align: center;
    }
    [data-testid="stSelectbox"] > div,
    div[data-baseweb="select"] > div,
    [data-testid="stSlider"] input[type="range"],
    [data-testid="stRadio"] label,
    button, select, [role="button"],
    [data-baseweb="select"] { cursor: pointer !important; }
    [data-testid="stAppViewContainer"] > [data-testid="stMain"] {
        transition: margin-left 0.3s ease;
    }
    </style>
    """, unsafe_allow_html=True)

    # =================================================================
    # WSPÓŁDZIELONA STANDARYZACJA (liczona raz)
    # Widma wzorcowe są najpierw UJEDNOLICANE na wspólną, rosnącą siatkę
    # osi X — dokładnie tą samą transformacją, przez którą przejdą nowe
    # widma w sekcji klasyfikacji. Dzięki temu identyczne wejście daje
    # identyczną reprezentację liczbową (self-match = 0), a k-NN wiernie
    # odtwarza podział ekspercki po ponownym wgraniu tego samego pliku.
    # =================================================================
    x_ref_ujedn, _przygotuj_widma = zbuduj_ujednolicacz_osi(x.values)
    macierz_wzorcowe_ujedn = _przygotuj_widma(krzywe.values, x.values)  # (n_widm, n_pkt)
    skaler_referencyjny = StandardScaler().fit(macierz_wzorcowe_ujedn)
    dane_std_oryginaly = skaler_referencyjny.transform(macierz_wzorcowe_ujedn)

    # =================================================================
    # SUGESTIA LICZBY KLASTRÓW (K)
    # =================================================================
    with st.expander("📐 Sugestia Optymalnej Liczby Klastrów (K)", expanded=False):
        st.markdown(
            "Wykresy pomagają dobrać właściwą liczbę grup **K** przed uruchomieniem "
            "klasteryzacji. Każda metoda patrzy na problem z innej strony."
        )
        if max_k < 3:
            st.warning("Za mało krzywych, aby sensownie porównać różne wartości K.")
        else:
            zakres_k, inercje, silhouettes, db_scores, calinski = oblicz_metryki_k(
                dane_std_oryginaly, 2, max_k
            )

            # Elbow: druga pochodna inercji (guard na krótkie zakresy)
            if len(inercje) >= 3:
                diff2 = np.diff(np.diff(inercje))
                k_elbow = zakres_k[int(np.argmax(diff2)) + 1]
            else:
                k_elbow = zakres_k[0]
            k_silhouette = zakres_k[int(np.argmax(silhouettes))]
            k_db = zakres_k[int(np.argmin(db_scores))]
            k_calinski = zakres_k[int(np.argmax(calinski))]

            # Kneedle — punkt maksymalnej odległości od prostej łączącej końce
            inercje_arr = np.array(inercje, dtype=float)
            x_norm = (np.array(zakres_k) - zakres_k[0]) / max(zakres_k[-1] - zakres_k[0], 1)
            y_norm = (inercje_arr - inercje_arr.min()) / (inercje_arr.max() - inercje_arr.min() + 1e-12)
            x0, y0 = x_norm[0], y_norm[0]
            x1, y1 = x_norm[-1], y_norm[-1]
            dx, dy = x1 - x0, y1 - y0
            odleglosci = np.abs(dy * x_norm - dx * y_norm + x1 * y0 - y1 * x0) / (np.sqrt(dx**2 + dy**2) + 1e-12)
            k_kneedle = zakres_k[int(np.argmax(odleglosci))]

            st.info(
                f"**Sugerowane K:** "
                f"Elbow (2. pochodna) \u2192 **{k_elbow}** | "
                f"Kneedle \u2192 **{k_kneedle}** | "
                f"Silhouette \u2192 **{k_silhouette}** | "
                f"Davies-Bouldin \u2192 **{k_db}** | "
                f"Calinski-Harabasz \u2192 **{k_calinski}**"
            )

            fig_k = make_subplots(
                rows=2, cols=3,
                subplot_titles=[
                    "Elbow (Inercja) \u2014 szukaj zgi\u0119cia",
                    "Kneedle \u2014 max odleg\u0142o\u015b\u0107 od prostej",
                    "Silhouette \u2014 im wy\u017cszy tym lepiej",
                    "Davies-Bouldin \u2014 im ni\u017cszy tym lepiej",
                    "Calinski-Harabasz \u2014 im wy\u017cszy tym lepiej",
                    "",
                ],
            )
            kolor_marker = "#d62728"

            def _trace(xk, yk, k_opt, kolor, row, col):
                fig_k.add_trace(go.Scatter(
                    x=xk, y=yk, mode="lines+markers",
                    line=dict(color=kolor, width=2),
                    marker=dict(
                        color=[kolor_marker if k == k_opt else kolor for k in xk],
                        size=[11 if k == k_opt else 7 for k in xk],
                    ),
                    hovertemplate="K=%{x}<br>=%{y:.4g}<extra></extra>",
                ), row=row, col=col)

            _trace(zakres_k, inercje,     k_elbow,      "#1f77b4", 1, 1)
            _trace(zakres_k, list(odleglosci), k_kneedle, "#8c564b", 1, 2)
            _trace(zakres_k, silhouettes, k_silhouette, "#2ca02c", 1, 3)
            _trace(zakres_k, db_scores,   k_db,         "#ff7f0e", 2, 1)
            _trace(zakres_k, calinski,    k_calinski,   "#9467bd", 2, 2)

            fig_k.update_layout(height=480, showlegend=False,
                                margin=dict(l=10, r=10, t=40, b=10))
            fig_k.update_xaxes(title_text="K", dtick=1)
            st.plotly_chart(fig_k, width='stretch')
            st.caption(
                "\U0001f534 Czerwony punkt = sugerowane K. "
                "Kneedle to geometryczna metoda wyznaczania 'kolana' krzywej inercji \u2014 "
                "szuka punktu o maksymalnej odleg\u0142o\u015bci od prostej \u0142\u0105cz\u0105cej kra\u0144ce."
            )

    with st.expander("Kompleksowy Opis Metodologiczny", expanded=False):
        c_d1, c_d2 = st.columns(2)
        with c_d1:
            st.markdown(f"#### Algorytm Główny: `{metoda}`")
            st.write(OPISY_METOD.get(metoda, ""))
        with c_d2:
            st.markdown(f"#### Obróbka Wstępna: `{optymalizacja}`")
            st.write(OPISY_PREPROCESSING.get(optymalizacja, ""))

    # =================================================================
    # KLASYFIKACJA REGUŁOWA WIDM EPR (Marciniak et al. 2025) — NOWE W v4
    # Deterministyczne drzewa decyzyjne z Fig. 1A/1B. Działa na widmach
    # wzorcowych po ujednoliceniu osi X (macierz_wzorcowe_ujedn na siatce
    # x_ref_ujedn), więc znaki f(g) i ekstrema są liczone spójnie z resztą
    # aplikacji. Nie wymaga uczenia ani Ground Truth.
    # =================================================================
    with st.expander("🧪 Klasyfikacja regułowa EPR (typy I–V wg Marciniak 2025)",
                     expanded=False):
        st.markdown(
            "Przypisuje każde widmo do jednego z pięciu typów line-shape "
            "(**I, II, III, IVA, IVB, V**) na podstawie drzew decyzyjnych z pracy "
            "*Marciniak et al. (2025), Front. Public Health 13:1659601* (Fig. 1). "
            "Metoda jest **deterministyczna i interpretowalna** — nie uczy się na "
            "danych, tylko sprawdza znak sygnału w wybranych wartościach g oraz "
            "obecność lokalnych ekstremów. Zwraca pełną ścieżkę decyzyjną."
        )
        st.caption(
            "Działa na tej samej ujednoliconej (rosnącej) osi g co reszta "
            "aplikacji. Zakłada preprocessing zbliżony do pracy (odjęcie tła "
            "rurki, liniowa korekcja bazy, normalizacja). Jeśli oś X jest w mT, "
            "przelicz ją najpierw na g-factor. To rekonstrukcja logiki drzew z "
            "rysunków pracy, nie oryginalny kod autorów — definicja ekstremum i "
            "sposób filtracji nie są w pracy w pełni sprecyzowane, dlatego "
            "wystawiono je jako parametry. Savitzky-Golay + detekcja przez "
            "pochodną są zwykle bliższe praktyce EPR niż średnia ruchoma + "
            "prominencja. Rozbieżności między metodami pojawiają się na widmach "
            "granicznych — obejrzyj je na wykresie i w tabeli zgodności poniżej."
        )

        # --- Wiersz 1: tryb + metody (wygładzanie, detekcja ekstremów) ---
        col_reg1, col_reg2, col_reg3 = st.columns(3)
        with col_reg1:
            reg_tryb = st.radio(
                "Rodzaj widm:",
                ["Nienapromienione (0 Gy)", "Napromienione (10 Gy)"],
                key="reg_tryb",
                help="Wybiera drzewo decyzyjne: Fig. 1A (0 Gy) lub 1B (10 Gy). "
                     "W trybie napromienionym typy IVA i IVB są nierozróżnialne.",
            )
        with col_reg2:
            reg_metoda_smooth = st.radio(
                "Metoda wygładzania:",
                ["Średnia ruchoma", "Savitzky-Golay"],
                key="reg_metoda_smooth",
                help="Średnia ruchoma jest prosta, ale spłaszcza i przesuwa "
                     "piki. Savitzky-Golay (wielomian 3. rzędu) zachowuje "
                     "wysokość i położenie pików — bliżej praktyki EPR i "
                     "pierwotnej metody.",
            )
        with col_reg3:
            reg_metoda_ekstr = st.radio(
                "Detekcja ekstremów:",
                ["Prominencja amplitudowa", "Zmiana znaku pochodnej"],
                key="reg_metoda_ekstr",
                help="Prominencja: ekstremum musi wystawać ponad brzegi okna. "
                     "Zmiana znaku pochodnej: klasyczny warunek min/max (pochodna "
                     "przechodzi przez zero) + filtr amplitudy. Różne definicje "
                     "mogą dać różny typ dla widm granicznych.",
            )

        # --- Wiersz 2: okno + prominencja ---
        col_reg4, col_reg5 = st.columns(2)
        with col_reg4:
            reg_okno = st.slider(
                "Okno wygładzania (liczba punktów):",
                min_value=3, max_value=21, value=7, step=2,
                key="reg_okno",
                help="Nieparzyste okno filtra wygładzającego przy wykrywaniu "
                     "lokalnych ekstremów (odsiewa szum).",
            )
        with col_reg5:
            reg_prom = st.number_input(
                "Próg prominencji (w jednostkach sygnału, a.u.):",
                min_value=0.0, value=0.0, format="%.4g", key="reg_prom",
                help="Bezwzględny próg amplitudy: ekstremum liczy się tylko, "
                     "gdy wystaje ponad brzegi okna o co najmniej tę wartość. "
                     "Jednostki są TAKIE SAME jak Twój sygnał (nie sigma!). "
                     "0 = brak filtra. Sugestia: wpisz ~3× szacowaną σ szumu "
                     "podaną poniżej.",
            )

        # --- Szacowanie σ szumu linii bazowej (skrzydło g ≈ 2.024–2.030) ---
        # Daje punkt odniesienia do progu prominencji ORAZ samo w sobie jest
        # informacją o jakości widma (poziom szumu tła).
        sigmy = []
        for i in range(n_krzywych):
            s = reg_szum_sigma(x_ref_ujedn, macierz_wzorcowe_ujedn[i],
                               smooth_window=reg_okno,
                               metoda_smooth=reg_metoda_smooth)
            if s is not None:
                sigmy.append(s)
        if sigmy:
            sig_med = float(np.median(sigmy))
            sig_min, sig_max = float(np.min(sigmy)), float(np.max(sigmy))
            st.caption(
                f"📉 Szacowana σ szumu linii bazowej (skrzydło g≈2.024–2.030): "
                f"mediana **{sig_med:.4g}** a.u. (zakres {sig_min:.4g}–{sig_max:.4g} "
                f"w zbiorze). Sugerowany próg prominencji ≈ 3σ = **{3*sig_med:.4g}** "
                f"a.u. Wyższa σ = bardziej zaszumione widmo."
            )
        else:
            st.caption(
                "📉 Nie udało się oszacować σ szumu — brak punktów w oknie "
                "g≈2.024–2.030 (sprawdź zakres osi g)."
            )

        reg_fn = (reg_klasyfikuj_nienapromienione
                  if reg_tryb.startswith("Nienapromienione")
                  else reg_klasyfikuj_napromienione)

        # Klasyfikacja wszystkich widm wzorcowych (na ujednoliconej siatce)
        reg_wiersze = []
        reg_typy = []
        for i, nazwa in enumerate(nazwy_krzywych):
            y_widmo = macierz_wzorcowe_ujedn[i]
            try:
                typ, sciezka = reg_fn(x_ref_ujedn, y_widmo,
                                      smooth_window=reg_okno,
                                      prominence=float(reg_prom),
                                      metoda_smooth=reg_metoda_smooth,
                                      metoda_ekstr=reg_metoda_ekstr)
                reg_typy.append(typ)
                reg_wiersze.append({
                    "Widmo": str(nazwa),
                    "Typ (line-shape)": typ,
                    "f(2.0000)": round(reg_value_at_g(x_ref_ujedn, y_widmo, 2.0000), 5),
                    "f(2.0171)": round(reg_value_at_g(x_ref_ujedn, y_widmo, 2.0171), 5),
                    "Ścieżka decyzyjna": " → ".join(sciezka),
                })
            except ValueError as e:
                reg_typy.append("BŁĄD")
                reg_wiersze.append({
                    "Widmo": str(nazwa),
                    "Typ (line-shape)": "BŁĄD",
                    "f(2.0000)": None,
                    "f(2.0171)": None,
                    "Ścieżka decyzyjna": str(e),
                })

        df_reg = pd.DataFrame(reg_wiersze)

        col_reg_tab, col_reg_wyk = st.columns([3, 2])
        with col_reg_tab:
            st.markdown("##### 📋 Przypisane typy:")
            st.dataframe(df_reg, hide_index=True, width='stretch')
            licznosci_typ = pd.Series(reg_typy).value_counts()
            st.caption("Liczność typów: " + ", ".join(
                f"{t}: {n}" for t, n in licznosci_typ.items()))

            # Rekomendacja dozymetryczna wg wniosków pracy (typy III i V)
            n_dozy = sum(1 for t in reg_typy if t in ("III", "V"))
            if n_dozy > 0:
                st.info(
                    f"💡 {n_dozy} widm(o) należy do typów **III/V**, które wg "
                    "Marciniak et al. (2025) są rekomendowane do dozymetrii "
                    "(odporność na UV i wygrzewanie BG bez utraty tła)."
                )

            csv_reg = df_reg.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button(
                "⬇️ Pobierz typy regułowe (CSV)",
                data=csv_reg,
                file_name=zbuduj_nazwe_eksportu(
                    "klasyfikacja_regulowa_epr", ".csv",
                    nazwa_pliku=nazwa_pliku_wzorcowego),
                mime="text/csv",
                width='stretch',
            )

        with col_reg_wyk:
            st.markdown("##### 📈 Widma z punktami decyzyjnymi:")
            reg_wybrane = st.multiselect(
                "Widma do podglądu:", [str(n) for n in nazwy_krzywych],
                default=[str(n) for n in nazwy_krzywych[:min(4, n_krzywych)]],
                key="reg_multiselect",
            )
            if reg_wybrane:
                fig_reg = go.Figure()
                mapa_nazw = {str(n): i for i, n in enumerate(nazwy_krzywych)}
                for nz in reg_wybrane:
                    i = mapa_nazw[nz]
                    fig_reg.add_trace(go.Scatter(
                        x=x_ref_ujedn, y=macierz_wzorcowe_ujedn[i],
                        mode="lines", name=f"{nz} ({reg_typy[i]})",
                        line=dict(width=1.4),
                    ))
                # Pionowe linie w punktach decyzyjnych g
                for g0 in _REG_PUNKTY_G:
                    fig_reg.add_vline(x=g0, line_dash="dot",
                                      line_color="gray", opacity=0.5)
                fig_reg.update_layout(
                    height=360, margin=dict(l=10, r=10, t=10, b=10),
                    # Oś g-factor malejąca w prawo — konwencja EPR
                    xaxis=dict(title="g-factor", showgrid=True,
                               gridcolor="#e0e0e0", autorange="reversed"),
                    yaxis=dict(title="Sygnał EPR (a.u.)", showgrid=True,
                               gridcolor="#e0e0e0"),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                bgcolor="rgba(255,255,255,0.8)"),
                    hovermode="closest",
                )
                st.plotly_chart(fig_reg, width='stretch')
                st.caption(
                    "Kropkowane pionowe linie: punkty g = 2.0000, 2.0043, 2.0171 "
                    "użyte w drzewie decyzyjnym."
                )

        # Macierz zgodności z podziałem eksperckim (jeśli dostępny GT)
        if gt_dostepny:
            st.markdown("---")
            st.markdown("##### 🔗 Zgodność typów regułowych z podziałem eksperckim:")
            df_krzyz = pd.crosstab(
                pd.Series([str(e) for e in etykiety_eksperta], name="Grupa ekspercka"),
                pd.Series(reg_typy, name="Typ regułowy"),
            )
            st.dataframe(df_krzyz, width='stretch')
            st.caption(
                "Tabela krzyżowa: ile widm z każdej grupy eksperckiej trafiło "
                "do poszczególnych typów regułowych. Pozwala sprawdzić, czy "
                "kategoryzacja z pracy pokrywa się z Twoim podziałem."
            )

    # =================================================================
    # PRZETWARZANIE DANYCH WEJŚCIOWYCH
    # =================================================================
    indeksy_oryginalow = None
    krzywe_aug = None

    if optymalizacja == "Augmentacja sygnału":
        krzywe_aug, etykiety_zrodlowe = augmentuj_sygnal(
            krzywe, aug_technika, aug_sila, aug_kopie,
            random_state=KONFIG["RANDOM_STATE"]
        )
        dane_do_algorytmu = StandardScaler().fit_transform(krzywe_aug.T)
        indeksy_oryginalow = list(range(n_krzywych))
    elif optymalizacja == "Analiza trendu":
        dane_do_algorytmu = StandardScaler().fit_transform(
            krzywe.diff(axis=0).fillna(0).T
        )
    elif optymalizacja == "UMAP (Redukcja topologiczna)" and umap_dostepne:
        dane_do_algorytmu = umap.UMAP(
            n_neighbors=KONFIG["UMAP_N_NEIGHBORS"],
            min_dist=KONFIG["UMAP_MIN_DIST_PREPROC"],
            random_state=KONFIG["RANDOM_STATE"],
        ).fit_transform(dane_std_oryginaly)
    elif optymalizacja == "FeatureExtraction":
        cechy = pd.DataFrame(index=nazwy_krzywych)
        cechy["Max"] = krzywe.max().values
        # Indeks df jest zresetowany (pozycyjny) — idxmax odpowiada pozycjom w x
        cechy["Poz_Max"] = krzywe.idxmax().apply(lambda idx: x.iloc[int(idx)]).values
        cechy["Srednia"] = krzywe.mean().values
        cechy["Std"] = krzywe.std().values
        cechy["Skosnosc"] = krzywe.skew().values
        cechy["Kurtoza"] = krzywe.kurt().values
        fft_amplitudy = np.abs(np.fft.rfft(krzywe, axis=0))
        for f_idx in range(1, min(4, fft_amplitudy.shape[0])):
            cechy[f"FFT_Skladowa_{f_idx}"] = fft_amplitudy[f_idx, :]
        dwt_a_mean, dwt_d_energy, dwt_d_std = [], [], []
        for col in krzywe.columns:
            sig = krzywe[col].values
            if len(sig) % 2 != 0:
                sig = sig[:-1]
            approx = (sig[0::2] + sig[1::2]) / np.sqrt(2)
            detail = (sig[0::2] - sig[1::2]) / np.sqrt(2)
            dwt_a_mean.append(np.mean(approx))
            dwt_d_energy.append(np.sum(detail ** 2))
            dwt_d_std.append(np.std(detail))
        cechy["DWT_Haar_A_Srednia"] = dwt_a_mean
        cechy["DWT_Haar_D_Energia"] = dwt_d_energy
        cechy["DWT_Haar_D_Std"] = dwt_d_std
        dane_do_algorytmu = StandardScaler().fit_transform(cechy)
    elif optymalizacja == "MinMaxScaler":
        dane_do_algorytmu = MinMaxScaler().fit_transform(krzywe.T)
    elif optymalizacja == "Filtrowanie szumów":
        dane_do_algorytmu = StandardScaler().fit_transform(
            krzywe.rolling(window=KONFIG["ROLLING_WINDOW"],
                           center=True, min_periods=1).mean().T
        )
    else:
        dane_do_algorytmu = dane_std_oryginaly

    # =================================================================
    # KLASTERYZACJA + METRYKI
    # =================================================================
    if optymalizacja == "Augmentacja sygnału":
        numery_grup = uruchom_silnik_klastrowania(
            metoda, dane_do_algorytmu, liczba_grup, liczba_grup,
            df_sygnaly_raw=krzywe_aug,
        )
        numery_grup = np.asarray(numery_grup)
        numery_grup_oryg = numery_grup[indeksy_oryginalow]
        krzywe_do_wykresu = krzywe
        nazwy_do_wykresu = nazwy_krzywych
        numery_grup_do_wykresu = numery_grup_oryg
    else:
        numery_grup = np.asarray(uruchom_silnik_klastrowania(
            metoda, dane_do_algorytmu, liczba_grup, liczba_grup,
            df_sygnaly_raw=krzywe,
        ))
        krzywe_do_wykresu = krzywe
        nazwy_do_wykresu = nazwy_krzywych
        numery_grup_do_wykresu = numery_grup

    # =================================================================
    # NAZWY KLASTRÓW LITERAMI GRUP EKSPERCKICH (a -> „Klaster A”)
    # Gdy dostępny jest podział ekspercki, każdemu klastrowi nadajemy
    # nazwę wg dominującej w nim grupy eksperckiej (przypisanie optymalne,
    # algorytm węgierski) wraz z procentem czystości. Bez Ground Truth
    # pozostają nazwy numeryczne („Klaster 1”).
    # =================================================================
    if gt_dostepny:
        mapa_liter = mapuj_klastry_na_litery(
            numery_grup_do_wykresu, etykiety_eksperta)
    else:
        mapa_liter = {}

    def nazwa_klastra(k_id):
        """Etykieta wyświetlana dla klastra (litera+czystość albo numer)."""
        k_id = int(k_id)
        if k_id in mapa_liter:
            return mapa_liter[k_id][0]
        return "Szum / Odrzuty" if k_id <= 0 else f"Klaster {k_id}"

    def nazwa_klastra_pelna(k_id):
        """Etykieta z procentem czystości, np. „Klaster A (82%)”."""
        k_id = int(k_id)
        if k_id in mapa_liter:
            etyk, czyst = mapa_liter[k_id]
            if czyst is not None:
                return f"{etyk} ({czyst:.0f}%)"
            return etyk
        return "Szum / Odrzuty" if k_id <= 0 else f"Klaster {k_id}"

    ari_score = None
    nmi_score = None
    if gt_dostepny:
        ari_score = adjusted_rand_score(etykiety_eksperta, numery_grup_do_wykresu) * 100
        nmi_score = normalized_mutual_info_score(etykiety_eksperta, numery_grup_do_wykresu) * 100

        st.markdown("### Skuteczność dopasowania:")
        kpi_ari, kpi_nmi = st.columns(2)
        kpi_ari.metric("Indeks ARI", f"{ari_score:.1f}%")
        kpi_nmi.metric("Indeks NMI", f"{nmi_score:.1f}%")
        if zrodlo_gt == "domyslny":
            st.caption(
                "⚠️ ARI/NMI liczone względem **domyślnego** podziału y1–y43 "
                "(brak arkusza Ground Truth w pliku)."
            )
    else:
        st.markdown("### Skuteczność dopasowania:")
        st.warning(
            "Brak Ground Truth — metryki zewnętrzne (ARI/NMI) niedostępne. "
            "Dodaj do pliku arkusz **Ground Truth** (kolumny: Krzywa | Grupa), "
            "aby je odblokować."
        )

    # =================================================================
    # WYKRES 1: WSZYSTKIE KRZYWE
    # =================================================================
    st.subheader("Wykres 1: Wszystkie sklasterowane krzywe")

    czy_dendrogram = "Hierarchiczna" in metoda and "+" not in metoda

    if czy_dendrogram:
        fig_dend, ax_dend = plt.subplots(figsize=(10, 4.2))
        if optymalizacja == "Augmentacja sygnału":
            etykiety_dendro = list(krzywe_aug.columns)
        else:
            etykiety_dendro = [str(n) for n in nazwy_do_wykresu]
        dendrogram(
            linkage(dane_do_algorytmu,
                    method="ward" if "Warda" in metoda else "average"),
            labels=etykiety_dendro, leaf_rotation=90, ax=ax_dend,
        )
        st.pyplot(fig_dend)
        plt.close(fig_dend)
    else:
        fig1 = go.Figure()
        dodane_do_legendy = set()
        for i, col in enumerate(krzywe_do_wykresu.columns):
            klaster_id = int(numery_grup_do_wykresu[i])
            if klaster_id > 0:
                kolor = PLOTLY_KOLORY[(klaster_id - 1) % 10]
                etykieta_grupy = nazwa_klastra(klaster_id)
            else:
                kolor = "#aaaaaa"
                etykieta_grupy = "Szum / Odrzuty"
            fig1.add_trace(go.Scatter(
                x=x,
                y=krzywe_do_wykresu[col],
                mode="lines",
                name=etykieta_grupy,
                legendgroup=etykieta_grupy,
                showlegend=(klaster_id not in dodane_do_legendy),
                line=dict(color=kolor, width=1.2),
                opacity=0.6,
                hovertemplate=(
                    f"<b>{col}</b><br>{etykieta_grupy}"
                    "<br>X: %{x}<br>Y: %{y:.4f}<extra></extra>"
                ),
            ))
            dodane_do_legendy.add(klaster_id)
        fig1.update_layout(
            height=420,
            margin=dict(l=10, r=10, t=10, b=10),
            legend=dict(groupclick="toggleitem",
                        bgcolor="rgba(255,255,255,0.8)", borderwidth=1),
            # Oś g-factor malejąca w prawo — konwencja EPR (pole rośnie w prawo)
            xaxis=dict(showgrid=True, gridcolor="#e0e0e0", autorange="reversed"),
            yaxis=dict(showgrid=True, gridcolor="#e0e0e0"),
            hovermode="closest",
        )
        st.plotly_chart(fig1, width='stretch')

    # =================================================================
    # WYKRES 2: PROFILE MODELOWE (ŚREDNIE + CIEŃ WARIANCJI)
    # =================================================================
    st.subheader("Wykres 2: Uśrednione profile modelowe (Wzorce kształtu fali)")

    fig2 = go.Figure()
    unikalne_klastry = sorted(set(int(v) for v in numery_grup_do_wykresu))

    for k_id in unikalne_klastry:
        maska_klastra = [int(numery_grup_do_wykresu[idx]) == k_id
                         for idx in range(len(numery_grup_do_wykresu))]
        krzywe_klastra = krzywe_do_wykresu.iloc[:, maska_klastra]
        if krzywe_klastra.shape[1] == 0:
            continue

        profil_sredni = krzywe_klastra.mean(axis=1)
        profil_std = krzywe_klastra.std(axis=1).fillna(0)
        gorna = profil_sredni + profil_std
        dolna = profil_sredni - profil_std

        if k_id > 0:
            kolor = PLOTLY_KOLORY[(k_id - 1) % 10]
            label_sredni = ("Wzorzec " + nazwa_klastra(k_id)
                            if k_id in mapa_liter
                            else f"Wzorzec Klastra {k_id}")
        else:
            kolor = "#aaaaaa"
            label_sredni = "Średnia Szumu"
        liczba_krzywych_kl = krzywe_klastra.shape[1]

        fig2.add_trace(go.Scatter(
            x=list(x) + list(x[::-1]),
            y=list(gorna) + list(dolna[::-1]),
            fill="toself",
            fillcolor=kolor,
            opacity=0.12,
            line=dict(width=0),
            showlegend=False,
            hoverinfo="skip",
            legendgroup=label_sredni,
        ))
        fig2.add_trace(go.Scatter(
            x=x,
            y=profil_sredni,
            mode="lines",
            name=label_sredni,
            legendgroup=label_sredni,
            line=dict(color=kolor, width=2.5),
            hovertemplate=(
                f"<b>{label_sredni}</b><br>"
                f"Liczba krzywych: {liczba_krzywych_kl}<br>"
                "X: %{x}<br>Średnia: %{y:.4f}<extra></extra>"
            ),
        ))

    fig2.update_layout(
        height=420,
        margin=dict(l=10, r=10, t=10, b=10),
        legend=dict(bgcolor="rgba(255,255,255,0.8)", borderwidth=1),
        # Oś g-factor malejąca w prawo — konwencja EPR
        xaxis=dict(showgrid=True, gridcolor="#e0e0e0", autorange="reversed"),
        yaxis=dict(showgrid=True, gridcolor="#e0e0e0"),
        hovermode="closest",
    )
    st.plotly_chart(fig2, width='stretch')

    # =================================================================
    # SKŁAD KLASTRÓW + EKSPORT
    # =================================================================
    klastry_slownik = {}
    for i, col in enumerate(krzywe_do_wykresu.columns):
        k_id = int(numery_grup_do_wykresu[i])
        klastry_slownik.setdefault(k_id, []).append(str(col))

    posortowane_klastry = sorted(klastry_slownik.keys())
    liczba_klastrow = len(posortowane_klastry)

    wiersze_eksportu = []
    for k_id in posortowane_klastry:
        if k_id == 0:
            nazwa_kl = "Szum / Odrzuty"
            nazwa_koloru = "Szary"
            czystosc_kl = None
        else:
            nazwa_kl = nazwa_klastra(k_id)
            nazwa_koloru = NAZWY_KOLOROW[(k_id - 1) % 10]
            czystosc_kl = (mapa_liter[k_id][1]
                           if k_id in mapa_liter else None)
        for krzywa in klastry_slownik[k_id]:
            wiersz = {
                "Krzywa": krzywa,
                "Klaster": nazwa_kl,
                "Kolor": nazwa_koloru,
                "Nr Klastra": k_id,
            }
            if gt_dostepny:
                wiersz["Czystość (%)"] = czystosc_kl
            wiersze_eksportu.append(wiersz)
    df_eksport_klastry = pd.DataFrame(wiersze_eksportu)

    # POPRAWKA #1: przycisk w sidebarze dodawany PO obliczeniach,
    # w tym samym rerunie — bez opóźnienia i bez martwego klucza
    with st.sidebar:
        st.markdown("---")
        st.caption("💾 Pobierz skład klastrów:")
        csv_bytes_sb = df_eksport_klastry.to_csv(
            index=False, encoding="utf-8-sig"
        ).encode("utf-8-sig")
        st.download_button(
            "⬇️ Pobierz CSV",
            data=csv_bytes_sb,
            file_name=zbuduj_nazwe_eksportu(
                "klastry", ".csv", nazwa_pliku=nazwa_pliku_wzorcowego,
                metoda=metoda, obrobka=optymalizacja, k=liczba_grup),
            mime="text/csv",
            width='stretch',
        )

    with st.expander("📊 Szczegółowy skład wygenerowanych klastrów", expanded=False):
        if liczba_klastrow > 0:
            kolumny_klastrow = st.columns(min(liczba_klastrow, 4))
            for idx, k_id in enumerate(posortowane_klastry):
                col_ui = kolumny_klastrow[idx % 4]
                with col_ui:
                    if k_id == 0:
                        st.markdown("**⚪ Szum / Odrzuty**")
                    else:
                        n_koloru = NAZWY_KOLOROW[(k_id - 1) % 10]
                        st.markdown(
                            f"**🔹 {nazwa_klastra_pelna(k_id)}** ({n_koloru})")
                    st.caption(f"Liczba: {len(klastry_slownik[k_id])}")
                    st.code(", ".join(klastry_slownik[k_id]), language="text")

        st.markdown("---")
        st.markdown("##### 💾 Pobierz wygenerowany skład klastrów:")
        col_exp_csv, col_exp_xlsx = st.columns(2)
        with col_exp_csv:
            csv_bytes = df_eksport_klastry.to_csv(
                index=False, encoding="utf-8-sig"
            ).encode("utf-8-sig")
            st.download_button(
                label="⬇️ Pobierz CSV",
                data=csv_bytes,
                file_name=zbuduj_nazwe_eksportu(
                    "sklady_klastrow", ".csv",
                    nazwa_pliku=nazwa_pliku_wzorcowego, metoda=metoda,
                    obrobka=optymalizacja, k=liczba_grup),
                mime="text/csv",
                width='stretch',
            )
        with col_exp_xlsx:
            bufor_xlsx = io.BytesIO()
            with pd.ExcelWriter(bufor_xlsx, engine="openpyxl") as writer:
                df_eksport_klastry.to_excel(writer, index=False,
                                            sheet_name="Skład Klastrów")
                arkusz = writer.sheets["Skład Klastrów"]
                for kol, szer in zip("ABCDE", (20, 22, 18, 12, 12)):
                    arkusz.column_dimensions[kol].width = szer

                x_vals = np.asarray(x, dtype=float)

                # --- Arkusze z DANYMI wykresów (X + serie Y do edycji) ---
                # Surowe serie: łatwo zmienić kolory/linie, usuwać/dodawać
                # serie i budować własny wykres w Excelu.
                try:
                    df_xy_krzywe = _df_krzywe_xy(
                        x_vals, krzywe_do_wykresu, numery_grup_do_wykresu,
                        nazwa_klastra)
                    _zapisz_df_z_szerokoscia(
                        writer, df_xy_krzywe, "Dane — krzywe")

                    df_xy_profile = _df_profile_xy(
                        x_vals, krzywe_do_wykresu, numery_grup_do_wykresu,
                        nazwa_klastra)
                    ark_prof = _zapisz_df_z_szerokoscia(
                        writer, df_xy_profile, "Dane — profile")
                    # Natywny wykres profili: rysujemy kolumny „— średnia”
                    # (co trzecia, począwszy od kol. 2), ciągłe, w kolorach grup.
                    try:
                        chart_p = LineChart()
                        chart_p.title = "Uśrednione profile modelowe"
                        chart_p.height = 10; chart_p.width = 18
                        chart_p.x_axis.title = "Oś X"; chart_p.y_axis.title = "Sygnał"
                        chart_p.x_axis.delete = False; chart_p.y_axis.delete = False
                        # Oś g-factor malejąca w prawo — konwencja EPR
                        chart_p.x_axis.scaling.orientation = "maxMin"
                        cats_p = Reference(ark_prof, min_col=1, min_row=2,
                                           max_row=1 + len(df_xy_profile))
                        kolumny_srednie = list(range(2, df_xy_profile.shape[1] + 1, 3))
                        grupy_prof = sorted(set(int(v)
                                                for v in numery_grup_do_wykresu))
                        for poz, kol_idx in enumerate(kolumny_srednie):
                            ref = Reference(ark_prof, min_col=kol_idx,
                                            max_col=kol_idx, min_row=1,
                                            max_row=1 + len(df_xy_profile))
                            chart_p.add_data(ref, titles_from_data=True)
                        chart_p.set_categories(cats_p)
                        for poz, seria in enumerate(chart_p.series):
                            gid = grupy_prof[poz] if poz < len(grupy_prof) else 0
                            kolor = ("AAAAAA" if gid <= 0
                                     else _hex(PLOTLY_KOLORY[(gid - 1) % 10]))
                            seria.graphicalProperties = GraphicalProperties()
                            seria.graphicalProperties.line = LineProperties(
                                solidFill=kolor, w=28575)
                            seria.smooth = False
                        ark_prof.add_chart(
                            chart_p, f"A{len(df_xy_profile) + 4}")
                    except Exception:
                        pass
                except Exception:
                    pass

                # --- Arkusz z wykresami publikacyjnymi (PNG, 300 DPI) ---
                # Podgląd/gotowy rysunek; dane do edycji są w arkuszach wyżej.
                try:
                    wb = writer.book
                    ark_wyk = wb.create_sheet("Wykresy (podgląd)")
                    ark_wyk["A1"] = ("Podgląd w rozdzielczości 300 DPI. "
                                     "Dane źródłowe do samodzielnej edycji "
                                     "wykresów znajdują się w arkuszach "
                                     "„Dane — krzywe” i „Dane — profile”.")
                    wiersz_ankor = 3
                    png1 = _png_krzywe(x_vals, krzywe_do_wykresu,
                                       numery_grup_do_wykresu, PLOTLY_KOLORY)
                    img1 = XLImage(png1); img1.anchor = f"A{wiersz_ankor}"
                    ark_wyk.add_image(img1)
                    wiersz_ankor += 24
                    png2 = _png_profile(x_vals, krzywe_do_wykresu,
                                        numery_grup_do_wykresu, PLOTLY_KOLORY)
                    img2 = XLImage(png2); img2.anchor = f"A{wiersz_ankor}"
                    ark_wyk.add_image(img2)
                    wiersz_ankor += 24
                    if czy_dendrogram:
                        etyk_d = [str(n) for n in nazwy_do_wykresu]
                        png3 = _png_dendrogram(dane_do_algorytmu, metoda, etyk_d)
                        if png3 is not None:
                            img3 = XLImage(png3); img3.anchor = f"A{wiersz_ankor}"
                            ark_wyk.add_image(img3)
                except Exception:
                    pass
            bufor_xlsx.seek(0)
            st.download_button(
                label="⬇️ Pobierz Excel (dane + wykresy)",
                data=bufor_xlsx.getvalue(),
                file_name=zbuduj_nazwe_eksportu(
                    "sklady_klastrow", ".xlsx",
                    nazwa_pliku=nazwa_pliku_wzorcowego, metoda=metoda,
                    obrobka=optymalizacja, k=liczba_grup),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch',
            )

    # =================================================================
    # KLASYFIKACJA NOWYCH WIDM DO KATEGORII WZORCOWYCH (WERSJA 2.2)
    # =================================================================
    with st.expander("🆕 Klasyfikacja nowych widm do kategorii wzorcowych",
                     expanded=False):
        # Źródło kategorii: podział ekspercki, a bez niego — bieżące klastry
        if gt_dostepny:
            klasy_referencyjne = [str(e) for e in etykiety_eksperta]
            zrodlo_klas = "kategorie eksperckie (Ground Truth / panel boczny)"
        else:
            klasy_referencyjne = [
                f"Klaster {int(v)}" if int(v) > 0 else "Szum"
                for v in numery_grup_do_wykresu
            ]
            zrodlo_klas = "klastry z bieżącej analizy (brak Ground Truth)"

        st.markdown(
            "Wgraj **osobny skoroszyt Excel** z nowymi widmami w tym samym "
            "układzie co dane główne: pierwsza kolumna = oś X, kolejne kolumny "
            "= widma. Każde nowe widmo zostanie przypisane do jednej z "
            f"kategorii wzorcowych. Źródło kategorii: **{zrodlo_klas}**."
        )
        st.caption(
            "Klasyfikacja działa w przestrzeni standaryzowanych surowych "
            "krzywych (ten sam skaler co dane wzorcowe), niezależnie od "
            "wybranej obróbki wstępnej. Jeśli oś X nowych widm różni się od "
            "referencyjnej, widma są interpolowane liniowo na siatkę wzorcową."
        )

        plik_nowe = st.file_uploader(
            "Wgraj plik z nowymi widmami (.xlsx)",
            type=["xlsx"], key="uploader_nowe_widma",
        )

        if plik_nowe is not None:
            nazwa_pliku_nowych = plik_nowe.name
            try:
                df_nowe_raw = pd.read_excel(plik_nowe, sheet_name=0, header=None)
                df_nowe = inteligentne_pobranie_tabeli(df_nowe_raw)
                x_nowe = df_nowe.iloc[:, 0].values.astype(float)
                widma_nowe = df_nowe.iloc[:, 1:]
                nazwy_nowe = [str(c) for c in widma_nowe.columns]

                if widma_nowe.shape[1] == 0:
                    st.error("W pliku nie znaleziono kolumn z widmami "
                             "(poza kolumną osi X).")
                else:
                    # Zabezpieczenie: zakresy osi X muszą się pokrywać,
                    # inaczej interpolacja zwróciłaby wartości brzegowe (płaskie).
                    if (x_nowe.min() > float(x_ref_ujedn.max())
                            or x_nowe.max() < float(x_ref_ujedn.min())):
                        st.error(
                            "Zakres osi X nowych widm nie pokrywa się z "
                            "referencyjnym — interpolacja niemożliwa. "
                            f"Nowe: [{x_nowe.min():.4g}, {x_nowe.max():.4g}], "
                            f"referencyjne: [{float(x_ref_ujedn.min()):.4g}, "
                            f"{float(x_ref_ujedn.max()):.4g}]. Sprawdź jednostki "
                            "osi X."
                        )
                        st.stop()

                    # UJEDNOLICENIE nową ścieżką: dokładnie ta sama funkcja,
                    # przez którą przeszły widma wzorcowe. Gwarantuje, że
                    # identyczne wejście = identyczna reprezentacja (self-match=0).
                    trzeba_interpolowac = (
                        len(x_nowe) != len(x_ref_ujedn)
                        or not np.allclose(np.sort(x_nowe), x_ref_ujedn)
                    )
                    if trzeba_interpolowac:
                        st.info(
                            f"Oś X nowych widm ({len(x_nowe)} pkt, zakres "
                            f"[{x_nowe.min():.4g}, {x_nowe.max():.4g}]) różni się "
                            f"od referencyjnej ({len(x_ref_ujedn)} pkt) — "
                            "zastosowano interpolację liniową na siatkę wzorcową."
                        )

                    # (n_widm, n_pkt) na wspólnej, rosnącej siatce referencyjnej
                    dane_nowe_ujedn = _przygotuj_widma(
                        widma_nowe.values.astype(float), x_nowe
                    )
                    # TEN SAM skaler co dane wzorcowe — kluczowe dla spójności
                    dane_nowe_std = skaler_referencyjny.transform(dane_nowe_ujedn)

                    metoda_klasyfikacji = st.radio(
                        "Metoda przypisania:",
                        ["k-NN (k=3, ważone odległością)",
                         "Najbliższy centroid kategorii",
                         "Skalibrowane prawdopodobieństwo (SVM + kalibracja)"],
                        horizontal=False, key="radio_klasyfikacja",
                    )

                    if "k-NN" in metoda_klasyfikacji:
                        k_snn = min(3, n_krzywych)
                        model_kl = KNeighborsClassifier(
                            n_neighbors=k_snn, weights="distance"
                        )
                        model_kl.fit(dane_std_oryginaly, klasy_referencyjne)
                        przypisania = model_kl.predict(dane_nowe_std)
                        proby = model_kl.predict_proba(dane_nowe_std)
                        pewnosci = proby.max(axis=1) * 100
                        opis_pewnosci = ("Zgodność głosów = udział głosów "
                                         "(ważonych odległością) zwycięskiej "
                                         "kategorii wśród k=3 najbliższych widm "
                                         "wzorcowych. To NIE jest kalibrowane "
                                         "prawdopodobieństwo.")
                        nazwa_kol_pewnosci = "Zgodność głosów (%)"

                    elif "Skalibrowane" in metoda_klasyfikacji:
                        # Skalibrowany klasyfikator probabilistyczny.
                        # Liczba foldów CV nie może przekroczyć liczności
                        # najmniejszej kategorii — inaczej kalibracja się wysypie.
                        licznosci = Counter(klasy_referencyjne)
                        min_klasa = min(licznosci.values())
                        n_klas = len(licznosci)

                        if min_klasa < 2 or n_klas < 2:
                            st.error(
                                "Kalibracja wymaga co najmniej 2 kategorii i "
                                "min. 2 widm wzorcowych w każdej. Najmniejsza "
                                f"kategoria ma tylko {min_klasa} widm(o). "
                                "Wybierz inną metodę lub dodaj więcej wzorców."
                            )
                            st.stop()

                        cv_folds = min(5, min_klasa)
                        # Bazowy SVM z RBF; sigmoidalna kalibracja Platta jest
                        # stabilniejsza od izotonicznej przy małych zbiorach.
                        baza_svm = SVC(kernel="rbf", probability=False,
                                       random_state=KONFIG["RANDOM_STATE"])
                        model_kl = CalibratedClassifierCV(
                            baza_svm, method="sigmoid", cv=cv_folds
                        )
                        model_kl.fit(dane_std_oryginaly, klasy_referencyjne)
                        przypisania = model_kl.predict(dane_nowe_std)
                        proby = model_kl.predict_proba(dane_nowe_std)
                        pewnosci = proby.max(axis=1) * 100
                        opis_pewnosci = (
                            f"Skalibrowane prawdopodobieństwo (SVM RBF + "
                            f"kalibracja Platta, {cv_folds}-krotna walidacja "
                            f"krzyżowa). Wartość przybliża rzeczywiste P(kategoria|"
                            f"widmo): 90% oznacza, że wśród wielu takich "
                            f"przypadków ok. 90% powinno należeć do wskazanej "
                            f"kategorii."
                        )
                        nazwa_kol_pewnosci = "Prawdopodobieństwo (%)"
                        if min_klasa < 5:
                            st.warning(
                                f"⚠️ Najmniejsza kategoria ma tylko {min_klasa} "
                                f"widm — kalibracja oparta na {cv_folds} foldach "
                                "może być niestabilna. Traktuj prawdopodobieństwa "
                                "orientacyjnie i zweryfikuj na wykresie."
                            )

                    else:
                        model_kl = NearestCentroid()
                        model_kl.fit(dane_std_oryginaly, klasy_referencyjne)
                        przypisania = model_kl.predict(dane_nowe_std)
                        centroidy = model_kl.centroids_
                        d = np.linalg.norm(
                            dane_nowe_std[:, None, :] - centroidy[None, :, :],
                            axis=2,
                        )
                        d_sort = np.sort(d, axis=1)
                        if d_sort.shape[1] >= 2:
                            pewnosci = (1.0 - d_sort[:, 0]
                                        / (d_sort[:, 1] + 1e-12)) * 100
                        else:
                            pewnosci = np.full(len(przypisania), 100.0)
                        opis_pewnosci = ("Margines separacji: jak bardzo "
                                         "najbliższy centroid wygrywa z drugim w "
                                         "kolejności (0 = remis, 100 = pełna "
                                         "dominacja). To wskaźnik zaufania, NIE "
                                         "prawdopodobieństwo.")
                        nazwa_kol_pewnosci = "Margines separacji"

                    df_wyniki_kl = pd.DataFrame({
                        "Nowe widmo": nazwy_nowe,
                        "Przypisana kategoria": [str(p) for p in przypisania],
                        nazwa_kol_pewnosci: np.round(pewnosci, 1),
                    })

                    col_kl1, col_kl2 = st.columns([2, 3])
                    with col_kl1:
                        st.markdown("##### 📋 Wyniki przypisania:")
                        st.dataframe(
                            df_wyniki_kl.style
                            .format({nazwa_kol_pewnosci: "{:.1f}"})
                            .background_gradient(subset=[nazwa_kol_pewnosci],
                                                 cmap="RdYlGn",
                                                 vmin=0, vmax=100),
                            hide_index=True, width='stretch',
                        )
                        st.caption(opis_pewnosci)
                        niskie = df_wyniki_kl[df_wyniki_kl[nazwa_kol_pewnosci] < 50]
                        if not niskie.empty:
                            st.warning(
                                f"⚠️ {len(niskie)} widm(o) z wartością < 50 — "
                                "przypisanie niepewne, zweryfikuj wzrokowo "
                                "na wykresie obok."
                            )

                        # „UCIEKINIERZY”: gdy wgrano ten sam plik co wzorcowy,
                        # a program dysponuje etykietami eksperckimi — pokaż
                        # widma, których przypisanie różni się od etykiety a–e.
                        if gt_dostepny:
                            mapa_ekspert = {
                                str(nz): str(et)
                                for nz, et in zip(nazwy_krzywych, etykiety_eksperta)
                            }
                            wiersze_uciek = []
                            for nz, prz in zip(nazwy_nowe, przypisania):
                                oczek = mapa_ekspert.get(str(nz))
                                if oczek is not None and str(prz) != oczek:
                                    wiersze_uciek.append({
                                        "Widmo": nz,
                                        "Grupa ekspercka": oczek,
                                        "Przypisano": str(prz),
                                    })
                            if wiersze_uciek:
                                st.markdown(
                                    f"##### 🔀 Rozbieżności z podziałem eksperckim "
                                    f"({len(wiersze_uciek)} z {len(nazwy_nowe)}):"
                                )
                                st.caption(
                                    "Te widma trafiły do innej grupy, niż wskazuje "
                                    "etykieta eksperta — są geometrycznie graniczne "
                                    "(pokrywają się z listą „czarnych owiec” i "
                                    "anomalii MSE)."
                                )
                                st.dataframe(pd.DataFrame(wiersze_uciek),
                                             hide_index=True,
                                             width='stretch')
                            else:
                                st.success(
                                    "✅ Wszystkie widma trafiły do swoich "
                                    "eksperckich grup — pełna zgodność."
                                )

                        col_kl_csv, col_kl_xlsx = st.columns(2)
                        with col_kl_csv:
                            csv_kl = df_wyniki_kl.to_csv(
                                index=False, encoding="utf-8-sig"
                            ).encode("utf-8-sig")
                            st.download_button(
                                "⬇️ Pobierz przypisania (CSV)",
                                data=csv_kl,
                                file_name=zbuduj_nazwe_eksportu(
                                    "klasyfikacja_nowych_widm", ".csv",
                                    nazwa_pliku=nazwa_pliku_wzorcowego,
                                    metoda=metoda_klasyfikacji,
                                    dodatek=(_oczysc_do_nazwy(
                                        nazwa_pliku_nowych, maxlen=30)
                                        if nazwa_pliku_nowych else None)),
                                mime="text/csv",
                                width='stretch',
                            )
                        with col_kl_xlsx:
                            bufor_kl = io.BytesIO()
                            with pd.ExcelWriter(bufor_kl,
                                                engine="openpyxl") as writer_kl:
                                # 1) Tabela przypisań (widmo, kategoria, pewność)
                                _zapisz_df_z_szerokoscia(
                                    writer_kl, df_wyniki_kl,
                                    "Przypisania", szer=22)
                                # 2) Dane wykresu nakładkowego (X + serie Y)
                                #    + natywny, edytowalny wykres Excela w stylu
                                #    aplikacji (wzorce ciągłe, nowe przerywane).
                                try:
                                    df_kl_xy = _df_klasyfikacja_xy(
                                        x_ref_ujedn, macierz_wzorcowe_ujedn,
                                        klasy_referencyjne, dane_nowe_ujedn,
                                        nazwy_nowe, przypisania)
                                    ark_dane_kl = _zapisz_df_z_szerokoscia(
                                        writer_kl, df_kl_xy, "Dane — wykres")
                                    # Kolory i style serii w kolejności kolumn:
                                    # najpierw wzorce (ciągłe), potem nowe (dash).
                                    kategorie_kl = sorted(
                                        set(str(k) for k in klasy_referencyjne))
                                    mapa_kol_kl = {
                                        kat: PLOTLY_KOLORY[i % 10]
                                        for i, kat in enumerate(kategorie_kl)
                                    }
                                    kolory_serii = [
                                        _hex(mapa_kol_kl[kat])
                                        for kat in kategorie_kl
                                    ]
                                    style_serii = ["solid"] * len(kategorie_kl)
                                    for prz in przypisania:
                                        kol = mapa_kol_kl.get(str(prz), "#aaaaaa")
                                        kolory_serii.append(_hex(kol))
                                        style_serii.append("dash")
                                    wykres_kl = _wykres_excel_klasyfikacja(
                                        ark_dane_kl, len(df_kl_xy),
                                        kolory_serii, style_serii)
                                    # Wykres pod danymi (2 wiersze odstępu) —
                                    # zawsze widoczny niezależnie od liczby serii.
                                    ark_dane_kl.add_chart(
                                        wykres_kl, f"A{len(df_kl_xy) + 4}")
                                except Exception:
                                    pass
                                # 3) Podgląd wykresu (PNG 300 DPI)
                                try:
                                    png_kl = _png_klasyfikacja(
                                        x_ref_ujedn, macierz_wzorcowe_ujedn,
                                        klasy_referencyjne, dane_nowe_ujedn,
                                        nazwy_nowe, przypisania, PLOTLY_KOLORY)
                                    if png_kl is not None:
                                        wb_kl = writer_kl.book
                                        ark_kl = wb_kl.create_sheet(
                                            "Wykres (podgląd)")
                                        ark_kl["A1"] = (
                                            "Podgląd 300 DPI. Dane źródłowe do "
                                            "edycji wykresu są w arkuszu "
                                            "„Dane — wykres”.")
                                        img_kl = XLImage(png_kl)
                                        img_kl.anchor = "A3"
                                        ark_kl.add_image(img_kl)
                                except Exception:
                                    pass
                            bufor_kl.seek(0)
                            st.download_button(
                                "⬇️ Pobierz Excel (dane + wykres)",
                                data=bufor_kl.getvalue(),
                                file_name=zbuduj_nazwe_eksportu(
                                    "klasyfikacja_nowych_widm", ".xlsx",
                                    nazwa_pliku=nazwa_pliku_wzorcowego,
                                    metoda=metoda_klasyfikacji,
                                    dodatek=(_oczysc_do_nazwy(
                                        nazwa_pliku_nowych, maxlen=30)
                                        if nazwa_pliku_nowych else None)),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                width='stretch',
                            )

                    with col_kl2:
                        st.markdown("##### 📈 Nowe widma na tle wzorców kategorii:")
                        unikalne_kategorie = sorted(set(klasy_referencyjne))
                        mapa_kolorow = {
                            kat: PLOTLY_KOLORY[i % 10]
                            for i, kat in enumerate(unikalne_kategorie)
                        }
                        fig_kl = go.Figure()
                        # Wzorcowe profile i nowe widma rysujemy na wspólnej,
                        # ujednoliconej (rosnącej) siatce osi X — spójnej z tym,
                        # na czym faktycznie działa klasyfikator.
                        klasy_ref_arr = np.asarray(
                            [str(k) for k in klasy_referencyjne])
                        # Grube linie: średnie profile kategorii wzorcowych
                        for kat in unikalne_kategorie:
                            maska_kat = klasy_ref_arr == kat
                            profil_kat = macierz_wzorcowe_ujedn[maska_kat].mean(axis=0)
                            fig_kl.add_trace(go.Scatter(
                                x=x_ref_ujedn, y=profil_kat, mode="lines",
                                name=f"Wzorzec: {kat}",
                                legendgroup=f"wzorzec_{kat}",
                                line=dict(color=mapa_kolorow[kat], width=3),
                            ))
                        # Przerywane linie: nowe widma w kolorze przypisania
                        for j, nazwa in enumerate(nazwy_nowe):
                            kat_j = str(przypisania[j])
                            fig_kl.add_trace(go.Scatter(
                                x=x_ref_ujedn, y=dane_nowe_ujedn[j], mode="lines",
                                name=f"{nazwa} → {kat_j}",
                                line=dict(color=mapa_kolorow.get(kat_j, "#aaaaaa"),
                                          width=1.5, dash="dash"),
                                opacity=0.85,
                                hovertemplate=(
                                    f"<b>{nazwa}</b> → {kat_j} "
                                    f"({pewnosci[j]:.0f}%)<br>"
                                    "X: %{x}<br>Y: %{y:.4f}<extra></extra>"
                                ),
                            ))
                        fig_kl.update_layout(
                            height=420,
                            margin=dict(l=10, r=10, t=10, b=10),
                            legend=dict(bgcolor="rgba(255,255,255,0.8)",
                                        borderwidth=1),
                            # Oś g-factor malejąca w prawo — konwencja EPR
                            xaxis=dict(showgrid=True, gridcolor="#e0e0e0",
                                       autorange="reversed"),
                            yaxis=dict(showgrid=True, gridcolor="#e0e0e0"),
                            hovermode="closest",
                        )
                        st.plotly_chart(fig_kl, width='stretch')

            except Exception as blad_kl:
                st.error(f"Nie udało się przetworzyć pliku z nowymi widmami: "
                         f"{blad_kl}")

    # =================================================================
    # MSE ANOMALY DETECTION
    # =================================================================
    with st.expander("🔍 Detekcja Anomalii MSE: Odległość od centroidu klastra",
                     expanded=False):
        st.markdown(
            "Dla każdej krzywej obliczana jest **fizyczna odległość MSE** od wzorca "
            "(centroidu) jej klastra. Krzywe z MSE powyżej progu `μ + 2σ` są "
            "automatycznie oznaczane jako anomalie."
        )

        wyniki_mse = []
        unikalne_k = sorted(set(int(v) for v in numery_grup_do_wykresu))
        for k_id in unikalne_k:
            if k_id == 0:
                continue  # pomijamy szum HDBSCAN
            maska_k = np.array([int(v) for v in numery_grup_do_wykresu]) == k_id
            dane_klastra = dane_std_oryginaly[maska_k]
            centroid = dane_klastra.mean(axis=0)
            for i, (nalezy, col) in enumerate(zip(maska_k, krzywe_do_wykresu.columns)):
                if nalezy:
                    mse = float(np.mean((dane_std_oryginaly[i] - centroid) ** 2))
                    wyniki_mse.append({
                        "Krzywa": str(col),
                        "Klaster": k_id,
                        "MSE od centroidu": round(mse, 4),
                    })

        df_mse = pd.DataFrame(wyniki_mse)
        if not df_mse.empty:
            prog_anomalii = (df_mse["MSE od centroidu"].mean()
                             + 2 * df_mse["MSE od centroidu"].std())
            df_mse["Anomalia"] = df_mse["MSE od centroidu"].apply(
                lambda v: "🚨 TAK" if v > prog_anomalii else "✅ NIE"
            )
            df_mse_sorted = df_mse.sort_values(
                "MSE od centroidu", ascending=False
            ).reset_index(drop=True)

            col_mse1, col_mse2 = st.columns(2)
            with col_mse1:
                st.markdown("##### 🚨 Anomalie (MSE > μ + 2σ):")
                anomalie = df_mse_sorted[
                    df_mse_sorted["Anomalia"] == "🚨 TAK"
                ].reset_index(drop=True)
                if not anomalie.empty:
                    st.dataframe(
                        anomalie[["Krzywa", "Klaster", "MSE od centroidu"]]
                        .style.format({"MSE od centroidu": "{:.4f}"})
                        .background_gradient(subset=["MSE od centroidu"], cmap="Reds"),
                        hide_index=True, width='stretch',
                    )
                    st.caption(
                        f"Próg anomalii: {prog_anomalii:.4f}  |  "
                        f"Wykryto: {len(anomalie)} krzywych"
                    )
                else:
                    st.success("Brak anomalii — wszystkie krzywe leżą blisko "
                               "centroidów klastrów.")
            with col_mse2:
                st.markdown("##### 📊 Ranking MSE wszystkich krzywych:")
                st.dataframe(
                    df_mse_sorted[["Krzywa", "Klaster", "MSE od centroidu", "Anomalia"]]
                    .style.format({"MSE od centroidu": "{:.4f}"}),
                    hide_index=True, width='stretch', height=320,
                )
        else:
            st.info("Brak klastrów do analizy (same odrzuty?).")

    # =================================================================
    # LEAVE-ONE-OUT — teraz TYLKO na żądanie + cache (POPRAWKA #9)
    # =================================================================
    with st.expander("🔬 Silnik Diagnostyczny Leave-One-Out (ARI / NMI)",
                     expanded=False):
        if not gt_dostepny:
            st.info("Analiza Leave-One-Out wymaga Ground Truth (ARI/NMI).")
        else:
            st.markdown(
                "Algorytm izoluje po kolei każdą krzywą z bazy danych, uruchamia "
                "grupowanie od nowa i bada, jak jej brak wpływa na globalny wskaźnik "
                "ARI. **Czarne Owce** — usunięcie podnosi wynik. "
                "**Filary Modelu** — usunięcie obniża wynik."
            )
            st.caption(
                f"⏱️ Analiza uruchamia **{n_krzywych}** pełnych klasteryzacji — "
                "dlatego startuje dopiero po kliknięciu przycisku. Wynik jest "
                "cache'owany dla bieżących ustawień."
            )

            if st.button("▶️ Uruchom analizę Leave-One-Out", key="btn_loo"):
                st.session_state["loo_uruchomione"] = True

            if st.session_state.get("loo_uruchomione", False):
                wyniki_loo = oblicz_leave_one_out(
                    metoda, dane_std_oryginaly, tuple(etykiety_eksperta),
                    krzywe, tuple(str(n) for n in nazwy_do_wykresu),
                    liczba_grup, ari_score,
                )
                df_loo = pd.DataFrame(wyniki_loo).sort_values(
                    by="Wpływ na ARI", ascending=False
                ).reset_index(drop=True)

                col_loo1, col_loo2 = st.columns(2)
                fmt_loo = {
                    "ARI bez krzywej (%)": "{:.2f}%",
                    "NMI bez krzywej (%)": "{:.2f}%",
                }
                with col_loo1:
                    st.markdown('##### 🚨 "Czarne Owce" (usunięcie PODNOSI ARI):')
                    df_czarne = df_loo[df_loo["Wpływ na ARI"] > 0.01].reset_index(drop=True)
                    if not df_czarne.empty:
                        st.dataframe(
                            df_czarne.style.format(
                                {**fmt_loo, "Wpływ na ARI": "+{:.2f}%"}
                            ),
                            hide_index=True, width='stretch',
                        )
                    else:
                        st.info("Brak wyraźnych anomalii. Wszystkie krzywe "
                                "wspierają model.")
                with col_loo2:
                    st.markdown('##### 🧱 "Filary Modelu" (usunięcie drastycznie '
                                'OBNIŻA ARI):')
                    df_filary = df_loo[df_loo["Wpływ na ARI"] < -0.01].sort_values(
                        "Wpływ na ARI"
                    ).reset_index(drop=True)
                    if not df_filary.empty:
                        st.dataframe(
                            df_filary.style.format(
                                {**fmt_loo, "Wpływ na ARI": "{:.2f}%"}
                            ),
                            hide_index=True, width='stretch',
                        )
                    else:
                        st.info("Brak kluczowych filarów — podział grup jest stabilny.")

    # =================================================================
    # RANKING SKUTECZNOŚCI ALGORYTMÓW
    # =================================================================
    with st.expander("🏆 Ranking Skuteczności Algorytmów", expanded=False):
        opis_metryk = ("ARI, NMI oraz Silhouette Score" if gt_dostepny
                       else "Silhouette Score (brak Ground Truth — bez ARI/NMI)")
        st.markdown(
            f"Ranking uruchamia każdą metodę **wielokrotnie** (na różnych "
            f"ziarnach losowych) i porównuje {opis_metryk} jako **średnią ± "
            f"odchylenie (σ)**. Kolumna **σ** i **Rozrzut σ** mówią o "
            f"STABILNOŚCI: metoda z wysokim wynikiem, ale dużym σ jest mniej "
            f"godna zaufania niż nieco słabsza, lecz powtarzalna. Metody "
            f"deterministyczne (Ward, korelacyjna, PCA+Ward) mają σ≈0 z "
            f"definicji — to zaleta, nie brak. Wyniki są cachowane."
        )
        st.caption(
            "⚠️ Dlaczego to ważne: na małym zbiorze pojedynczy wynik ARI potrafi "
            "być dziełem przypadku (szczęśliwe ziarno). Dopiero rozrzut po wielu "
            "ziarnach pokazuje, czy przewaga metody jest realna, czy losowa. "
            "Jeśli dobierasz metodę pod ten sam podział ekspercki, traktuj wysoki "
            "wynik ostrożnie — to nie jest już niezależny test."
        )
        st.caption(
            "🛡️ Ochrona pamięci: metody deterministyczne (Ward, PCA+Ward, "
            "korelacyjna, NMF) liczone są raz (σ=0 z definicji), a ciężkie "
            "(UMAP, Spectral, SOM) mają ograniczoną liczbę ziaren — dlatego w "
            "kolumnie Udane pow. zobaczysz różne mianowniki. To celowe: chroni "
            "przed przekroczeniem pamięci na Streamlit Cloud."
        )

        METODY_SZYBKIE = [
            "Hierarchiczna Aglomeracyjna (metoda Warda)",
            "PCA + Hierarchiczna (metoda Warda)",
            "K-means",
            "GMM (Probabilistyczna)",
            "Hierarchiczna Korelacyjna (metoda średnich)",
            "Spectral Clustering",
            "NMF (Nieujemna Faktoryzacja Macierzy)",
            "BGMM (Bayesowski GMM)",
        ]

        col_tr, col_zi = st.columns([2, 1])
        with col_tr:
            tryb_rankingu = st.radio(
                "Zakres rankingu:",
                ["⚡ Szybki (8 metod)", "🔬 Pełny (wszystkie metody)"],
                horizontal=True,
                key="tryb_rankingu",
            )
        with col_zi:
            n_ziaren_rank = st.slider(
                "Liczba ziaren (powtórzeń):", min_value=1, max_value=15,
                value=5, key="n_ziaren_rank",
                help="Ile razy uruchomić każdą metodę na różnych ziarnach "
                     "losowych. Więcej ziaren = pewniejsza ocena stabilności, "
                     "ale dłuższe liczenie. Metody deterministyczne liczą się "
                     "raz, a ciężkie (UMAP/Spectral/SOM) mają limit 3 ziaren "
                     "dla ochrony pamięci — niezależnie od tego suwaka.",
            )
        lista_do_rankingu = (METODY_SZYBKIE if "Szybki" in tryb_rankingu
                             else lista_metod)

        krzywe_dla_rankingu = krzywe_aug if krzywe_aug is not None else krzywe
        rekordy, bledy_rankingu = oblicz_ranking(
            dane_do_algorytmu,
            tuple(etykiety_eksperta) if gt_dostepny else None,
            krzywe_dla_rankingu,
            tuple(lista_do_rankingu),
            liczba_grup,
            tuple(indeksy_oryginalow) if indeksy_oryginalow is not None else None,
            n_ziaren=n_ziaren_rank,
        )

        if bledy_rankingu:
            with st.expander("⚠️ Metody pominięte w rankingu", expanded=False):
                for b in bledy_rankingu:
                    st.caption(f"• {b}")

        if rekordy:
            df_lb = pd.DataFrame(rekordy).sort_values(
                "Średnia (%)", ascending=False
            ).reset_index(drop=True)

            metryki = [c for c in ["ARI (%)", "NMI (%)", "Silhouette (%)"]
                       if c in df_lb.columns]

            klucz_sel = f"ranking_selekcja_{tryb_rankingu}_{liczba_grup}_{n_ziaren_rank}"
            if (klucz_sel not in st.session_state
                    or len(st.session_state[klucz_sel]) != len(df_lb)):
                st.session_state[klucz_sel] = [False] * len(df_lb)

            df_lb.insert(0, "Wybierz", st.session_state[klucz_sel])

            st.caption("Zaznacz metody które chcesz porównać, następnie kliknij "
                       "**Porównaj**. Kolumny **σ** to odchylenie po ziarnach "
                       "(mniej = stabilniej). **Rozrzut σ** = łączna niestabilność.")

            konfiguracja_kolumn = {
                "Wybierz": st.column_config.CheckboxColumn("✔", width="small"),
                "Algorytm AI": st.column_config.TextColumn("Algorytm AI", disabled=True),
                "Średnia (%)": st.column_config.NumberColumn("Średnia (%)",
                                                             disabled=True,
                                                             format="%.2f"),
            }
            # Metryki główne + ich odchylenia
            for m in metryki:
                konfiguracja_kolumn[m] = st.column_config.NumberColumn(
                    m, disabled=True, format="%.2f"
                )
                kol_sigma = m.replace(" (%)", " σ")
                if kol_sigma in df_lb.columns:
                    konfiguracja_kolumn[kol_sigma] = st.column_config.NumberColumn(
                        kol_sigma, disabled=True, format="%.2f",
                        help="Odchylenie standardowe tej metryki po ziarnach "
                             "losowych. Mniejsze = bardziej powtarzalne.",
                    )
            if "Rozrzut σ" in df_lb.columns:
                konfiguracja_kolumn["Rozrzut σ"] = st.column_config.NumberColumn(
                    "Rozrzut σ", disabled=True, format="%.2f",
                    help="Łączna niestabilność metody (pierwiastek ze średniej "
                         "wariancji metryk). Im mniej, tym pewniejszy wynik.",
                )
            if "Udane pow." in df_lb.columns:
                konfiguracja_kolumn["Udane pow."] = st.column_config.TextColumn(
                    "Udane pow.", disabled=True,
                    help="Ile ziaren dało poprawny podział (≥2 klastry).",
                )

            df_edytowalny = st.data_editor(
                df_lb,
                hide_index=True,
                width='stretch',
                column_config=konfiguracja_kolumn,
                key=f"ranking_editor_{klucz_sel}",
            )

            st.session_state[klucz_sel] = df_edytowalny["Wybierz"].tolist()
            zaznaczone = df_edytowalny[df_edytowalny["Wybierz"] == True]
            wszystkie_zaznaczone = len(zaznaczone) == len(df_edytowalny)

            col_btn1, col_btn2, col_info = st.columns([1, 1, 3])
            with col_btn1:
                if st.button(
                    "☑️ Odznacz wszystkie" if wszystkie_zaznaczone
                    else "✅ Wybierz wszystkie",
                    width='stretch',
                    key="btn_wybierz_wszystkie",
                ):
                    st.session_state[klucz_sel] = [not wszystkie_zaznaczone] * len(df_lb)
                    st.rerun()
            with col_btn2:
                porownaj = st.button(
                    "📊 Porównaj zaznaczone",
                    width='stretch',
                    disabled=len(zaznaczone) < 2,
                    key="btn_porownaj",
                )
            with col_info:
                if len(zaznaczone) < 2:
                    st.caption("⬅️ Zaznacz co najmniej 2 metody żeby porównać.")
                else:
                    st.caption(f"Zaznaczono **{len(zaznaczone)}** metod do porównania.")

            if porownaj or st.session_state.get("ranking_porownanie_aktywne", False):
                if porownaj:
                    st.session_state["ranking_porownanie_aktywne"] = True
                    st.session_state["ranking_porownanie_df"] = zaznaczone.drop(
                        columns=["Wybierz"]
                    ).reset_index(drop=True)

                df_por = st.session_state.get("ranking_porownanie_df", pd.DataFrame())
                if not df_por.empty:
                    st.markdown("---")
                    st.markdown("#### 📊 Porównanie wybranych metod")

                    metryki_por = [c for c in ["ARI (%)", "NMI (%)", "Silhouette (%)"]
                                   if c in df_por.columns]
                    # Mapowanie metryka -> kolumna odchylenia (dla error bars)
                    sigma_kol = {"ARI (%)": "ARI σ", "NMI (%)": "NMI σ",
                                 "Silhouette (%)": "Silhouette σ"}
                    fig_por = go.Figure()
                    for idx, row_por in df_por.iterrows():
                        bledy_y = [float(row_por.get(sigma_kol.get(m, ""), 0) or 0)
                                   for m in metryki_por]
                        fig_por.add_trace(go.Bar(
                            name=row_por["Algorytm AI"],
                            x=metryki_por,
                            y=[row_por[m] for m in metryki_por],
                            error_y=dict(type="data", array=bledy_y, visible=True,
                                         thickness=1.3, width=4),
                            marker_color=PLOTLY_KOLORY[idx % 10],
                            text=[f"{row_por[m]:.1f}" for m in metryki_por],
                            textposition="outside",
                        ))
                    fig_por.update_layout(
                        barmode="group",
                        height=380,
                        margin=dict(l=10, r=10, t=30, b=10),
                        legend=dict(orientation="h", yanchor="bottom", y=1.02),
                        yaxis=dict(range=[0, 115], title="Wartość (%) ± σ"),
                        xaxis=dict(title="Metryka"),
                    )
                    st.plotly_chart(fig_por, width='stretch')
                    st.caption(
                        "Wąsy błędów = odchylenie standardowe po ziarnach. "
                        "Nakładające się wąsy dwóch metod oznaczają, że ich "
                        "różnica może być nieistotna (mieści się w rozrzucie losowym)."
                    )

                    df_styl = df_por.set_index("Algorytm AI")

                    def podswietl_max(s):
                        return ["background-color: #d4edda; font-weight: bold"
                                if v == s.max() else "" for v in s]

                    fmt_por = {m: "{:.2f}" for m in metryki_por + ["Średnia (%)"]}
                    for kol in ["ARI σ", "NMI σ", "Silhouette σ", "Rozrzut σ"]:
                        if kol in df_styl.columns:
                            fmt_por[kol] = "{:.2f}"
                    st.dataframe(
                        df_styl.style
                        .apply(podswietl_max, subset=metryki_por)
                        .format(fmt_por),
                        width='stretch',
                    )

                    najlepsza = df_por.loc[df_por["Średnia (%)"].idxmax(),
                                           "Algorytm AI"]
                    st.success(f"🏆 Najwyższa średnia w porównaniu: **{najlepsza}**")
                    # Najstabilniejsza = najmniejszy Rozrzut σ (jeśli dostępny)
                    if "Rozrzut σ" in df_por.columns and len(df_por) > 1:
                        najstab = df_por.loc[df_por["Rozrzut σ"].idxmin(),
                                             "Algorytm AI"]
                        if najstab != najlepsza:
                            st.info(
                                f"🎯 Najstabilniejsza (najmniejszy rozrzut): "
                                f"**{najstab}**. Jeśli różnica średnich jest "
                                f"mniejsza niż σ, stabilność może być lepszym "
                                f"kryterium wyboru niż sam wynik."
                            )

                    if st.button("✖️ Zamknij porównanie", key="btn_zamknij_por"):
                        st.session_state["ranking_porownanie_aktywne"] = False
                        st.session_state["ranking_porownanie_df"] = pd.DataFrame()
                        st.rerun()
        else:
            st.info("Brak wyników — sprawdź dane wejściowe.")

except Exception as ob_blad:
    st.error(f"Błąd krytyczny podczas renderowania: {ob_blad}")
    if KONFIG["DEBUG"]:
        st.exception(ob_blad)  # pełny traceback do debugowania
